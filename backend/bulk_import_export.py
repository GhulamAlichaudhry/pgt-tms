"""
Bulk Import/Export Service
CSV/Excel import and export for all entities
"""
import csv
import io
from typing import List, Dict, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
import models

class BulkImportExportService:
    def __init__(self, db: Session):
        self.db = db
    
    # ============================================
    # IMPORT FUNCTIONS
    # ============================================
    
    def import_clients_csv(self, csv_file) -> Dict:
        """Import clients from CSV"""
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            for row in results['total'], row_data in enumerate(csv_reader, 1):
                results['total'] += 1
                try:
                    client = models.Client(
                        name=row_data['name'],
                        contact_person=row_data.get('contact_person'),
                        phone=row_data.get('phone'),
                        email=row_data.get('email'),
                        address=row_data.get('address'),
                        credit_limit=float(row_data.get('credit_limit', 0)),
                        payment_terms=int(row_data.get('payment_terms', 30))
                    )
                    self.db.add(client)
                    results['success'] += 1
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {results['total']}: {str(e)}")
            
            self.db.commit()
            
        except Exception as e:
            results['errors'].append(f"File error: {str(e)}")
        
        return results
    
    def import_vendors_csv(self, csv_file) -> Dict:
        """Import vendors from CSV"""
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            for row_data in csv_reader:
                results['total'] += 1
                try:
                    vendor = models.Vendor(
                        name=row_data['name'],
                        contact_person=row_data.get('contact_person'),
                        phone=row_data.get('phone'),
                        email=row_data.get('email'),
                        address=row_data.get('address'),
                        payment_terms=int(row_data.get('payment_terms', 30))
                    )
                    self.db.add(vendor)
                    results['success'] += 1
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {results['total']}: {str(e)}")
            
            self.db.commit()
            
        except Exception as e:
            results['errors'].append(f"File error: {str(e)}")
        
        return results
    
    def import_staff_csv(self, csv_file) -> Dict:
        """Import staff from CSV"""
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            for row_data in csv_reader:
                results['total'] += 1
                try:
                    staff = models.Staff(
                        employee_id=row_data['employee_id'],
                        name=row_data['name'],
                        position=row_data['position'],
                        gross_salary=float(row_data['gross_salary']),
                        monthly_deduction=float(row_data.get('monthly_deduction', 0))
                    )
                    self.db.add(staff)
                    results['success'] += 1
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {results['total']}: {str(e)}")
            
            self.db.commit()
            
        except Exception as e:
            results['errors'].append(f"File error: {str(e)}")
        
        return results
    
    def import_vehicles_csv(self, csv_file) -> Dict:
        """Import vehicles from CSV"""
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            for row_data in csv_reader:
                results['total'] += 1
                try:
                    vehicle = models.Vehicle(
                        vehicle_no=row_data['vehicle_no'],
                        vehicle_type=row_data['vehicle_type'],
                        capacity_tons=float(row_data['capacity_tons'])
                    )
                    self.db.add(vehicle)
                    results['success'] += 1
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {results['total']}: {str(e)}")
            
            self.db.commit()
            
        except Exception as e:
            results['errors'].append(f"File error: {str(e)}")
        
        return results
    
    # ============================================
    # EXPORT FUNCTIONS
    # ============================================
    
    def export_clients_excel(self) -> io.BytesIO:
        """Export all clients to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # Headers
        headers = ['ID', 'Client Code', 'Name', 'Contact Person', 'Phone', 'Email', 
                  'Address', 'Current Balance', 'Credit Limit', 'Payment Terms', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        clients = self.db.query(models.Client).all()
        for client in clients:
            ws.append([
                client.id,
                client.client_code or '',
                client.name,
                client.contact_person or '',
                client.phone or '',
                client.email or '',
                client.address or '',
                client.current_balance,
                client.credit_limit,
                client.payment_terms,
                'Active' if client.is_active else 'Inactive'
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_vendors_excel(self) -> io.BytesIO:
        """Export all vendors to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendors"
        
        # Headers
        headers = ['ID', 'Vendor Code', 'Name', 'Contact Person', 'Phone', 'Email', 
                  'Address', 'Current Balance', 'Payment Terms', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        vendors = self.db.query(models.Vendor).all()
        for vendor in vendors:
            ws.append([
                vendor.id,
                vendor.vendor_code or '',
                vendor.name,
                vendor.contact_person or '',
                vendor.phone or '',
                vendor.email or '',
                vendor.address or '',
                vendor.current_balance,
                vendor.payment_terms,
                'Active' if vendor.is_active else 'Inactive'
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_staff_excel(self) -> io.BytesIO:
        """Export all staff to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff"
        
        # Headers
        headers = ['ID', 'Employee ID', 'Name', 'Position', 'Gross Salary', 
                  'Advance Balance', 'Monthly Deduction', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        staff_list = self.db.query(models.Staff).all()
        for staff in staff_list:
            ws.append([
                staff.id,
                staff.employee_id,
                staff.name,
                staff.position,
                staff.gross_salary,
                staff.advance_balance,
                staff.monthly_deduction,
                'Active' if staff.is_active else 'Inactive'
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_trips_excel(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> io.BytesIO:
        """Export trips to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trips"
        
        # Headers
        headers = ['ID', 'Date', 'Reference No', 'Vehicle', 'Client', 'Vendor', 
                  'Source', 'Destination', 'Tonnage', 'Client Freight', 'Vendor Freight',
                  'Gross Profit', 'Net Profit', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        query = self.db.query(models.Trip)
        
        if start_date:
            query = query.filter(models.Trip.date >= start_date)
        if end_date:
            query = query.filter(models.Trip.date <= end_date)
        
        trips = query.all()
        
        for trip in trips:
            ws.append([
                trip.id,
                trip.date.strftime('%Y-%m-%d') if trip.date else '',
                trip.reference_no,
                trip.vehicle.vehicle_no if trip.vehicle else '',
                trip.client.name if trip.client else '',
                trip.vendor.name if trip.vendor else '',
                trip.source_location,
                trip.destination_location,
                trip.total_tonnage,
                trip.client_freight,
                trip.vendor_freight,
                trip.gross_profit,
                trip.net_profit,
                trip.status.value if hasattr(trip.status, 'value') else str(trip.status)
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    # ============================================
    # TEMPLATE GENERATION
    # ============================================
    
    def generate_import_template(self, entity_type: str) -> io.BytesIO:
        """Generate CSV template for import"""
        templates = {
            'clients': ['name', 'contact_person', 'phone', 'email', 'address', 'credit_limit', 'payment_terms'],
            'vendors': ['name', 'contact_person', 'phone', 'email', 'address', 'payment_terms'],
            'staff': ['employee_id', 'name', 'position', 'gross_salary', 'monthly_deduction'],
            'vehicles': ['vehicle_no', 'vehicle_type', 'capacity_tons']
        }
        
        if entity_type not in templates:
            raise ValueError(f"Unknown entity type: {entity_type}")
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(templates[entity_type])
        
        # Add sample row
        if entity_type == 'clients':
            writer.writerow(['ABC Company', 'John Doe', '+92-XXX-XXXXXXX', 'client@example.com', 'Address', '100000', '30'])
        elif entity_type == 'vendors':
            writer.writerow(['XYZ Vendor', 'Jane Smith', '+92-XXX-XXXXXXX', 'vendor@example.com', 'Address', '30'])
        elif entity_type == 'staff':
            writer.writerow(['EMP001', 'John Doe', 'Driver', '50000', '5000'])
        elif entity_type == 'vehicles':
            writer.writerow(['ABC-123', 'Truck', '10'])
        
        buffer.seek(0)
        return io.BytesIO(buffer.getvalue().encode('utf-8'))

# Singleton instance
def get_bulk_service(db: Session) -> BulkImportExportService:
    return BulkImportExportService(db)
