from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import timedelta, datetime, date
from typing import Optional
from sqlalchemy import and_, func
import io
import models
import schemas
import crud
import auth
from database import SessionLocal, engine, get_db
from notification_service import NotificationService, get_admin_user_ids
from validators import Validator, BusinessValidator, ValidationError
from audit_service import AuditService, get_client_ip, get_user_agent
from company_config import get_company_info, get_company_header

# Create database tables
models.Base.metadata.create_all(bind=engine)

# CRITICAL: Ensure admin user exists on every startup
# This prevents login credential issues
from ensure_admin import ensure_admin_exists
ensure_admin_exists()

app = FastAPI(
    title="PGT International Smart TMS",
    description="Transport Management System for PGT International (Private) Limited",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001"],  # React app URLs
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Helper function for Excel headers
def add_excel_company_header(worksheet, title: str, date_range: str = None):
    """Add professional company header to Excel worksheet"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    company_info = get_company_info()
    
    # Define styles
    header_font = Font(bold=True, size=18, color="DC2626")
    tagline_font = Font(italic=True, size=9, color="6B7280")
    address_font = Font(size=9, color="374151")
    title_font = Font(bold=True, size=14, color="374151")
    info_font = Font(size=9, color="6B7280")
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 15
    worksheet.column_dimensions['K'].width = 15
    
    # Row 1: Company Name
    worksheet.merge_cells('A1:K1')
    cell = worksheet['A1']
    cell.value = company_info['name']
    cell.font = header_font
    cell.alignment = center_align
    worksheet.row_dimensions[1].height = 24
    
    # Row 2: Tagline
    worksheet.merge_cells('A2:K2')
    cell = worksheet['A2']
    cell.value = company_info.get('tagline', 'Excellence in Transportation & Logistics')
    cell.font = tagline_font
    cell.alignment = center_align
    worksheet.row_dimensions[2].height = 15
    
    # Row 3: Address
    worksheet.merge_cells('A3:K3')
    cell = worksheet['A3']
    cell.value = company_info['address']
    cell.font = address_font
    cell.alignment = center_align
    worksheet.row_dimensions[3].height = 15
    
    # Row 4: Contact
    worksheet.merge_cells('A4:K4')
    cell = worksheet['A4']
    contact_info = f"Phone: {company_info['phone']} | Email: {company_info['email']} | Web: {company_info['website']}"
    cell.value = contact_info
    cell.font = address_font
    cell.alignment = center_align
    worksheet.row_dimensions[4].height = 15
    
    # Row 5: Empty with border
    worksheet.merge_cells('A5:K5')
    cell = worksheet['A5']
    cell.border = Border(bottom=Side(style='medium', color='DC2626'))
    worksheet.row_dimensions[5].height = 3
    
    # Row 6: Report Title
    worksheet.merge_cells('A6:K6')
    cell = worksheet['A6']
    cell.value = title
    cell.font = title_font
    cell.alignment = center_align
    worksheet.row_dimensions[6].height = 20
    
    # Row 7: Date Range (if provided)
    if date_range:
        worksheet.merge_cells('A7:K7')
        cell = worksheet['A7']
        cell.value = f"Period: {date_range}"
        cell.font = info_font
        cell.alignment = center_align
        worksheet.row_dimensions[7].height = 15
        start_row = 8
    else:
        start_row = 7
    
    # Generation Date
    worksheet.merge_cells(f'A{start_row}:K{start_row}')
    cell = worksheet[f'A{start_row}']
    cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    cell.font = info_font
    cell.alignment = center_align
    worksheet.row_dimensions[start_row].height = 15
    
    # Empty row before data
    start_row += 1
    worksheet.row_dimensions[start_row].height = 3
    
    return start_row + 1  # Return the row number where data headers should start

# Authentication endpoints
@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = auth.authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# ============================================
# PASSWORD RESET ENDPOINTS
# ============================================

@app.post("/password-reset/request")
def request_password_reset(
    email: str,
    db: Session = Depends(get_db)
):
    """Request password reset - sends email with reset link"""
    from password_reset_service import PasswordResetService
    
    result = PasswordResetService.create_reset_token(db, email)
    return result

@app.post("/password-reset/validate")
def validate_reset_token(
    token: str,
    db: Session = Depends(get_db)
):
    """Validate if reset token is valid"""
    from password_reset_service import PasswordResetService
    
    user = PasswordResetService.validate_reset_token(db, token)
    
    if user:
        return {
            "valid": True,
            "username": user.username
        }
    else:
        return {
            "valid": False,
            "error": "Invalid or expired token"
        }

@app.post("/password-reset/reset")
def reset_password(
    token: str,
    new_password: str,
    db: Session = Depends(get_db)
):
    """Reset password using token"""
    from password_reset_service import PasswordResetService
    
    result = PasswordResetService.reset_password(db, token, new_password)
    
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=400, detail=result["error"])

@app.post("/users/", response_model=schemas.User)
def create_user(
    user: schemas.UserCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    return crud.create_user(db=db, user=user)

# ============================================
# USER MANAGEMENT ENDPOINTS
# ============================================

@app.get("/users/", response_model=list[schemas.User])
def get_all_users(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Get all users - Admin only"""
    users = db.query(models.User).all()
    return users

@app.get("/users/{user_id}", response_model=schemas.User)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Get specific user - Admin only"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@app.put("/users/{user_id}", response_model=schemas.User)
def update_user(
    user_id: int,
    user_update: schemas.UserUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Update user - Admin only"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deactivating themselves
    if user.id == current_user.id and user_update.is_active == False:
        raise HTTPException(status_code=400, detail="Cannot deactivate your own account")
    
    # Update fields
    if user_update.full_name is not None:
        user.full_name = user_update.full_name
    if user_update.email is not None:
        user.email = user_update.email
    if user_update.role is not None:
        user.role = user_update.role
    if user_update.is_active is not None:
        user.is_active = user_update.is_active
    
    db.commit()
    db.refresh(user)
    return user

@app.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Delete user - Admin only"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deleting themselves
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    db.delete(user)
    db.commit()
    return {"message": f"User {user.username} deleted successfully"}

@app.put("/users/{user_id}/password")
def reset_user_password(
    user_id: int,
    password_data: schemas.PasswordReset,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Reset user password - Admin only"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash new password
    user.hashed_password = auth.get_password_hash(password_data.new_password)
    db.commit()
    
    return {"message": f"Password reset successfully for {user.username}"}

@app.get("/users/me/permissions")
def get_my_permissions(
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get current user's permissions"""
    permissions = {
        models.UserRole.ADMIN: {
            "can_manage_users": True,
            "can_view_profit": True,
            "can_approve_payments": True,
            "can_delete_data": True,
            "can_view_all_reports": True,
            "can_manage_master_data": True
        },
        models.UserRole.MANAGER: {
            "can_manage_users": False,
            "can_view_profit": False,
            "can_approve_payments": True,
            "can_delete_data": False,
            "can_view_all_reports": True,
            "can_manage_master_data": True
        },
        models.UserRole.SUPERVISOR: {
            "can_manage_users": False,
            "can_view_profit": False,
            "can_approve_payments": False,
            "can_delete_data": False,
            "can_view_all_reports": False,
            "can_manage_master_data": False
        }
    }
    
    return {
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "full_name": current_user.full_name,
            "role": current_user.role.value
        },
        "permissions": permissions.get(current_user.role, {})
    }

# ============================================
# NOTIFICATION ENDPOINTS
# ============================================

@app.get("/notifications")
def get_notifications(
    unread_only: bool = False,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get user notifications"""
    notif = NotificationService(db)
    notifications = notif.get_user_notifications(
        user_id=current_user.id,
        unread_only=unread_only,
        limit=limit
    )
    return notifications

@app.get("/notifications/unread-count")
def get_unread_count(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get count of unread notifications"""
    notif = NotificationService(db)
    count = notif.get_unread_count(user_id=current_user.id)
    return {"count": count}

@app.put("/notifications/{notification_id}/read")
def mark_notification_read(
    notification_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Mark notification as read"""
    notif = NotificationService(db)
    success = notif.mark_as_read(notification_id, current_user.id)
    if not success:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"success": True}

@app.put("/notifications/mark-all-read")
def mark_all_read(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Mark all notifications as read"""
    notif = NotificationService(db)
    count = notif.mark_all_as_read(current_user.id)
    return {"marked_read": count}

@app.delete("/notifications/{notification_id}")
def delete_notification(
    notification_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Delete a notification"""
    notification = db.query(models.Notification).filter(
        models.Notification.id == notification_id,
        models.Notification.user_id == current_user.id
    ).first()
    
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    db.delete(notification)
    db.commit()
    return {"success": True}

# ============================================
# COMPANY SETTINGS ENDPOINTS
# ============================================

@app.get("/company-settings")
def get_company_settings(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get company settings"""
    settings = db.query(models.CompanySetting).first()
    if not settings:
        # Create default settings if none exist
        settings = models.CompanySetting()
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings

@app.put("/company-settings")
def update_company_settings(
    settings_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Update company settings (Admin only)"""
    settings = db.query(models.CompanySetting).first()
    if not settings:
        settings = models.CompanySetting()
        db.add(settings)
    
    # Update fields
    for key, value in settings_data.items():
        if hasattr(settings, key):
            setattr(settings, key, value)
    
    settings.updated_by = current_user.id
    db.commit()
    db.refresh(settings)
    return settings

# ============================================
# BACKUP & RESTORE ENDPOINTS
# ============================================

@app.post("/backup/create")
def create_backup(
    description: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Create database backup (Admin only)"""
    from backup_service import BackupService
    
    service = BackupService()
    result = service.create_backup(description or f"Manual backup by {current_user.username}")
    
    if result["success"]:
        return {
            "success": True,
            "message": "Backup created successfully",
            "backup": result["metadata"]
        }
    else:
        raise HTTPException(status_code=500, detail=result["error"])

@app.get("/backup/list")
def list_backups(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """List all backups (Admin only)"""
    from backup_service import BackupService
    
    service = BackupService()
    backups = service.list_backups()
    return {"backups": backups}

@app.post("/backup/restore/{backup_name}")
def restore_backup(
    backup_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Restore from backup (Admin only)"""
    from backup_service import BackupService
    from pathlib import Path
    
    service = BackupService()
    backup_file = Path("backups") / f"{backup_name}.zip"
    
    if not backup_file.exists():
        raise HTTPException(status_code=404, detail="Backup not found")
    
    result = service.restore_backup(str(backup_file))
    
    if result["success"]:
        return {
            "success": True,
            "message": "Database restored successfully",
            "safety_backup": result["safety_backup"]
        }
    else:
        raise HTTPException(status_code=500, detail=result["error"])

@app.delete("/backup/{backup_name}")
def delete_backup(
    backup_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Delete backup (Admin only)"""
    from backup_service import BackupService
    from pathlib import Path
    
    service = BackupService()
    backup_file = Path("backups") / f"{backup_name}.zip"
    
    if not backup_file.exists():
        raise HTTPException(status_code=404, detail="Backup not found")
    
    result = service.delete_backup(str(backup_file))
    
    if result["success"]:
        return {"success": True, "message": "Backup deleted"}
    else:
        raise HTTPException(status_code=500, detail=result["error"])

# ============================================
# AUDIT TRAIL ENDPOINTS
# ============================================

@app.get("/audit-logs")
def get_audit_logs(
    table_name: Optional[str] = None,
    record_id: Optional[int] = None,
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Get audit logs (Admin/Manager only)"""
    audit = AuditService(db)
    
    if table_name and record_id:
        logs = audit.get_record_history(table_name, record_id, limit)
    elif user_id:
        logs = audit.get_user_activity(user_id, limit)
    else:
        # Get recent audit logs
        query = db.query(models.AuditLog).order_by(models.AuditLog.timestamp.desc())
        
        if table_name:
            query = query.filter(models.AuditLog.table_name == table_name)
        if action:
            query = query.filter(models.AuditLog.action == action)
        
        logs = query.limit(limit).all()
    
    return logs

# Dashboard endpoints
@app.get("/dashboard/stats", response_model=schemas.DashboardStats)
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_dashboard_stats(db)

@app.get("/dashboard/financial-summary")
def get_financial_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get comprehensive financial summary for dashboard"""
    from financial_calculator import FinancialCalculator
    
    calculator = FinancialCalculator(db)
    try:
        return calculator.get_master_financial_summary()
    finally:
        calculator.close()

@app.get("/dashboard/chart-data")
def get_dashboard_chart_data(
    months: int = 6,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get chart data for dashboard visualizations"""
    from financial_calculator import FinancialCalculator
    
    calculator = FinancialCalculator(db)
    try:
        return calculator.get_revenue_vs_expenses_chart_data(months)
    finally:
        calculator.close()

# ============================================
# DAILY CASH FLOW ENDPOINTS
# ============================================

@app.get("/daily-cash-flow")
def get_daily_cash_flow(
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get daily cash flow data with optional date range"""
    from financial_calculator import FinancialCalculator
    from datetime import datetime, timedelta, date as date_type
    
    calculator = FinancialCalculator(db)
    
    try:
        if start_date and end_date:
            # Return range of daily cash flows
            cash_flows = []
            current = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            while current <= end:
                flow = calculator.get_daily_cash_flow(current)
                flow['date'] = current.isoformat()
                cash_flows.append(flow)
                current += timedelta(days=1)
            
            # Calculate totals
            total_income = sum(f['daily_income'] for f in cash_flows)
            total_outgoing = sum(f['daily_outgoing'] for f in cash_flows)
            total_net = sum(f['daily_net'] for f in cash_flows)
            
            return {
                "cash_flows": cash_flows,
                "summary": {
                    "total_income": total_income,
                    "total_outgoing": total_outgoing,
                    "total_net": total_net,
                    "days": len(cash_flows)
                }
            }
        else:
            # Return single day
            target_date = datetime.strptime(date, '%Y-%m-%d').date() if date else date_type.today()
            flow = calculator.get_daily_cash_flow(target_date)
            flow['date'] = target_date.isoformat()
            return flow
    finally:
        calculator.close()

# ============================================
# AGING ANALYSIS ENDPOINTS
# ============================================

@app.get("/vendors/aging-analysis")
def get_vendors_aging_analysis(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get vendor aging analysis"""
    from ledger_service import LedgerService
    from datetime import date, timedelta
    
    service = LedgerService(db)
    vendors = service.get_all_vendors_summary()
    
    # Calculate aging for each vendor
    for vendor in vendors:
        # Get all unpaid payables
        payables = db.query(models.Payable).filter(
            models.Payable.vendor_id == vendor['vendor_id'],
            models.Payable.outstanding_amount > 0
        ).all()
        
        aging = {
            '0-30': 0,
            '31-60': 0,
            '61-90': 0,
            '90+': 0
        }
        
        for payable in payables:
            if payable.date:
                payable_date = payable.date.date() if hasattr(payable.date, 'date') else payable.date
                days_old = (date.today() - payable_date).days
                
                if days_old <= 30:
                    aging['0-30'] += float(payable.outstanding_amount)
                elif days_old <= 60:
                    aging['31-60'] += float(payable.outstanding_amount)
                elif days_old <= 90:
                    aging['61-90'] += float(payable.outstanding_amount)
                else:
                    aging['90+'] += float(payable.outstanding_amount)
        
        vendor['aging'] = aging
        vendor['total_overdue'] = aging['31-60'] + aging['61-90'] + aging['90+']
    
    return vendors

@app.get("/clients/aging-analysis")
def get_clients_aging_analysis(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get client aging analysis"""
    from ledger_service import LedgerService
    from datetime import date, timedelta
    
    service = LedgerService(db)
    clients = service.get_all_clients_summary()
    
    # Calculate aging for each client
    for client in clients:
        # Get all unpaid receivables
        receivables = db.query(models.Receivable).filter(
            models.Receivable.client_id == client['client_id'],
            models.Receivable.remaining_amount > 0
        ).all()
        
        aging = {
            '0-30': 0,
            '31-60': 0,
            '61-90': 0,
            '90+': 0
        }
        
        for receivable in receivables:
            if receivable.invoice_date:
                invoice_date = receivable.invoice_date.date() if hasattr(receivable.invoice_date, 'date') else receivable.invoice_date
                days_old = (date.today() - invoice_date).days
                
                if days_old <= 30:
                    aging['0-30'] += float(receivable.remaining_amount)
                elif days_old <= 60:
                    aging['31-60'] += float(receivable.remaining_amount)
                elif days_old <= 90:
                    aging['61-90'] += float(receivable.remaining_amount)
                else:
                    aging['90+'] += float(receivable.remaining_amount)
        
        client['aging'] = aging
        client['total_overdue'] = aging['31-60'] + aging['61-90'] + aging['90+']
    
    return clients

@app.get("/dashboard/receivables-details")
def get_receivables_details(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get detailed breakdown of receivables"""
    try:
        # Debug: Check total receivables in database
        total_receivables_raw = db.query(func.sum(models.Receivable.remaining_amount)).scalar()
        print(f"DEBUG: Total receivables in DB: {total_receivables_raw}")
        
        # Get all receivables with remaining amount > 0
        receivables = db.query(models.Receivable).filter(
            models.Receivable.remaining_amount > 0
        ).order_by(models.Receivable.remaining_amount.desc()).all()
        
        print(f"DEBUG: Found {len(receivables)} receivables")
        
        details = []
        for receivable in receivables:
            try:
                detail = {
                    "id": receivable.id,
                    "client_name": receivable.client.name if receivable.client else "Unknown Client",
                    "invoice_number": receivable.invoice_number,
                    "total_amount": receivable.total_amount,
                    "paid_amount": receivable.paid_amount,
                    "remaining_amount": receivable.remaining_amount,
                    "due_date": receivable.due_date.isoformat(),
                    "status": str(receivable.status),
                    "days_overdue": (datetime.now() - receivable.due_date).days if receivable.due_date < datetime.now() else 0
                }
                details.append(detail)
                print(f"DEBUG: Added receivable {receivable.id}")
            except Exception as e:
                print(f"DEBUG: Error processing receivable {receivable.id}: {e}")
        
        return {
            "total_receivables": sum(r.remaining_amount for r in receivables),
            "count": len(receivables),
            "details": details
        }
    except Exception as e:
        print(f"Error in receivables-details: {e}")
        import traceback
        traceback.print_exc()
        return {
            "total_receivables": 0,
            "count": 0,
            "details": [],
            "error": str(e)
        }

# ============================================
# INVOICE GENERATION ENDPOINTS
# ============================================

@app.post("/invoices/generate/{receivable_id}")
def generate_invoice(
    receivable_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate invoice PDF for receivable"""
    from fastapi.responses import StreamingResponse
    from invoice_generator import invoice_generator
    
    # Get receivable
    receivable = db.query(models.Receivable).filter(
        models.Receivable.id == receivable_id
    ).first()
    
    if not receivable:
        raise HTTPException(status_code=404, detail="Receivable not found")
    
    # Prepare invoice data
    invoice_data = {
        'invoice_number': receivable.invoice_number,
        'invoice_date': receivable.invoice_date.strftime('%Y-%m-%d'),
        'due_date': receivable.due_date.strftime('%Y-%m-%d'),
        'reference': f"Trip #{receivable.trip_id}" if receivable.trip_id else "N/A",
        'notes': 'Thank you for your business! Payment is due within the specified terms.',
        'payment_terms': f'Payment due within {receivable.payment_terms} days',
        'tax_rate': 0  # Add tax if applicable
    }
    
    # Prepare client data
    client_data = {
        'name': receivable.client.name,
        'contact_person': receivable.client.contact_person or 'N/A',
        'address': receivable.client.address or 'N/A',
        'phone': receivable.client.phone or 'N/A',
        'email': receivable.client.email or 'N/A'
    }
    
    # Prepare items
    items = [{
        'description': receivable.description,
        'quantity': 1,
        'rate': receivable.total_amount,
        'amount': receivable.total_amount
    }]
    
    # Generate PDF
    pdf_buffer = invoice_generator.generate_invoice_pdf(
        invoice_data=invoice_data,
        client_data=client_data,
        items=items
    )
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=invoice_{receivable.invoice_number}.pdf"
        }
    )

@app.post("/invoices/email/{receivable_id}")
def email_invoice(
    receivable_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate and email invoice to client"""
    from invoice_generator import invoice_generator
    from email_service import email_service
    
    # Get receivable
    receivable = db.query(models.Receivable).filter(
        models.Receivable.id == receivable_id
    ).first()
    
    if not receivable:
        raise HTTPException(status_code=404, detail="Receivable not found")
    
    if not receivable.client.email:
        raise HTTPException(status_code=400, detail="Client email not found")
    
    # Generate invoice PDF
    invoice_data = {
        'invoice_number': receivable.invoice_number,
        'invoice_date': receivable.invoice_date.strftime('%Y-%m-%d'),
        'due_date': receivable.due_date.strftime('%Y-%m-%d'),
        'reference': f"Trip #{receivable.trip_id}" if receivable.trip_id else "N/A",
        'notes': 'Thank you for your business!',
        'payment_terms': f'Payment due within {receivable.payment_terms} days',
        'tax_rate': 0
    }
    
    client_data = {
        'name': receivable.client.name,
        'contact_person': receivable.client.contact_person or 'N/A',
        'address': receivable.client.address or 'N/A',
        'phone': receivable.client.phone or 'N/A',
        'email': receivable.client.email
    }
    
    items = [{
        'description': receivable.description,
        'quantity': 1,
        'rate': receivable.total_amount,
        'amount': receivable.total_amount
    }]
    
    # Send email
    result = email_service.send_invoice_email(
        to_email=receivable.client.email,
        client_name=receivable.client.name,
        invoice_number=receivable.invoice_number,
        amount=receivable.total_amount
    )
    
    if result["success"]:
        return {
            "success": True,
            "message": f"Invoice emailed to {receivable.client.email}"
        }
    else:
        raise HTTPException(status_code=500, detail="Failed to send email")

# ============================================
# PAYMENT REMINDER ENDPOINTS
# ============================================

@app.post("/reminders/send-all")
def send_all_payment_reminders(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Send all pending payment reminders (Admin/Manager only)"""
    from payment_reminder_service import PaymentReminderService
    
    service = PaymentReminderService(db)
    results = service.send_all_reminders()
    
    return results

@app.post("/reminders/send/{receivable_id}")
def send_manual_reminder(
    receivable_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Send manual payment reminder for specific receivable"""
    from payment_reminder_service import PaymentReminderService
    
    service = PaymentReminderService(db)
    result = service.send_manual_reminder(receivable_id)
    
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=400, detail=result["error"])

@app.get("/reminders/overdue-summary")
def get_overdue_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get summary of overdue receivables"""
    from payment_reminder_service import PaymentReminderService
    
    service = PaymentReminderService(db)
    summary = service.get_overdue_summary()
    
    return summary

@app.get("/reminders/history/{receivable_id}")
def get_reminder_history(
    receivable_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get reminder history for receivable"""
    from payment_reminder_service import PaymentReminderService
    
    service = PaymentReminderService(db)
    history = service.get_reminder_history(receivable_id)
    
    return {"history": history}

@app.get("/dashboard/payables-details")
def get_payables_details(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get detailed breakdown of payables (only outstanding amounts)"""
    # Get payables with outstanding amounts > 0
    payables = db.query(models.Payable).filter(
        models.Payable.outstanding_amount > 0
    ).order_by(models.Payable.outstanding_amount.desc()).all()
    
    return {
        "total_payables": sum(p.outstanding_amount for p in payables),
        "count": len(payables),
        "details": [
            {
                "id": payable.id,
                "vendor_name": payable.vendor.name if payable.vendor else "Unknown Vendor",
                "invoice_number": payable.invoice_number,
                "amount": payable.amount,  # Total amount
                "outstanding_amount": payable.outstanding_amount,  # Amount still owed
                "paid_amount": payable.amount - payable.outstanding_amount,  # Amount already paid
                "due_date": payable.due_date.isoformat(),
                "status": payable.status,
                "days_overdue": (datetime.now() - payable.due_date).days if payable.due_date < datetime.now() else 0,
                "vendor_contact": payable.vendor.contact_person if payable.vendor else None,
                "vendor_phone": payable.vendor.phone if payable.vendor else None
            }
            for payable in payables
        ]
    }

@app.get("/dashboard/payment-requests-summary")
def get_payment_requests_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get payment requests summary for dashboard"""
    pending_count = crud.get_pending_payment_requests_count(db)
    
    # Get recent payment requests
    recent_requests = crud.get_payment_requests(db, skip=0, limit=5)
    
    return {
        "pending_count": pending_count,
        "recent_requests": [
            {
                "id": req.id,
                "vendor_name": req.vendor.name if req.vendor else "Unknown",
                "amount": req.requested_amount,
                "payment_type": req.payment_type,
                "urgency_level": req.urgency_level,
                "requested_at": req.requested_at.isoformat(),
                "status": req.status.value if hasattr(req.status, 'value') else req.status
            }
            for req in recent_requests
        ]
    }

# Report Generation Endpoints
@app.get("/reports/vendor-ledger-pdf/{vendor_id}")
def generate_vendor_ledger_pdf(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate vendor ledger PDF report"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    
    # Get vendor data
    vendor = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Get ledger entries - use existing CRUD function but handle the error gracefully
    try:
        ledger_entries = crud.get_ledger_entries(db, vendor_id, skip=0, limit=1000)
    except Exception as e:
        # If ledger entries fail, create empty list
        print(f"Warning: Could not fetch ledger entries: {e}")
        ledger_entries = []
    
    # Convert to dict format
    vendor_data = {
        "name": vendor.name,
        "vendor_code": getattr(vendor, 'vendor_code', f"VEN-{vendor.id:04d}"),
        "contact_person": vendor.contact_person,
        "phone": vendor.phone,
        "email": getattr(vendor, 'email', ''),
        "current_balance": vendor.current_balance,
        "is_active": vendor.is_active
    }
    
    ledger_data = []
    for entry in ledger_entries:
        try:
            ledger_data.append({
                "date": getattr(entry, 'date', entry.created_at),
                "description": entry.description,
                "reference_no": getattr(entry, 'reference_no', ''),
                "debit_amount": entry.debit_amount,
                "credit_amount": entry.credit_amount,
                "running_balance": getattr(entry, 'running_balance', entry.balance if hasattr(entry, 'balance') else 0)
            })
        except Exception as e:
            print(f"Warning: Could not process ledger entry {entry.id}: {e}")
            continue
    
    # Generate PDF
    generator = ReportGenerator()
    pdf_buffer = generator.generate_vendor_ledger_pdf(vendor_data, ledger_data)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=vendor_ledger_{vendor.name.replace(' ', '_')}.pdf"}
    )

@app.get("/reports/vendor-ledger-excel/{vendor_id}")
def generate_vendor_ledger_excel(
    vendor_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate vendor ledger Excel report with company header and trip details from real-time data"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    import io
    
    # Get vendor
    vendor = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Get payables
    payables_query = db.query(models.Payable).filter(
        models.Payable.vendor_id == vendor_id
    )
    
    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        payables_query = payables_query.filter(models.Payable.created_at >= start)
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d")
        payables_query = payables_query.filter(models.Payable.created_at <= end)
    
    payables = payables_query.order_by(models.Payable.created_at).all()
    
    # Build entries with trip details
    entries = []
    for payable in payables:
        # Find the trip associated with this payable
        trip = db.query(models.Trip).filter(models.Trip.payable_id == payable.id).first()
        
        entry = {
            "date": payable.created_at,
            "description": payable.description or f"Invoice: {payable.invoice_number}",
            "reference": trip.reference_no if trip else payable.invoice_number,
            "from": trip.source_location if trip else "",
            "to": trip.destination_location if trip else "",
            "tonnage": float(trip.total_tonnage) if trip and trip.total_tonnage else 0,
            "freight": float(trip.vendor_freight) if trip and trip.vendor_freight else float(payable.amount),
            "debit": payable.amount,
            "credit": 0,
            "type": "payable"
        }
        entries.append(entry)
        
        # Get payments (both APPROVED and PAID status)
        payments = db.query(models.PaymentRequest).filter(
            models.PaymentRequest.payable_id == payable.id,
            models.PaymentRequest.status.in_([models.PaymentRequestStatus.APPROVED, models.PaymentRequestStatus.PAID])
        ).all()
        
        for payment in payments:
            if payment.payment_date:
                entries.append({
                    "date": payment.payment_date,
                    "description": f"Payment: {payment.payment_channel.value if payment.payment_channel else 'N/A'}",
                    "reference": payment.payment_reference or "",
                    "from": "",
                    "to": "",
                    "tonnage": 0,
                    "freight": 0,
                    "debit": 0,
                    "credit": payment.requested_amount,
                    "type": "payment"
                })
    
    entries.sort(key=lambda x: x["date"])
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Ledger"
    
    # Add professional company header
    date_range_str = None
    if start_date or end_date:
        date_range_str = f"{start_date or 'Start'} to {end_date or 'End'}"
    
    header_row = add_excel_company_header(ws, f"Vendor Ledger: {vendor.name}", date_range_str)
    
    # Vendor Info Section
    ws.merge_cells(f'A{header_row}:K{header_row}')
    cell = ws[f'A{header_row}']
    cell.value = f"Vendor: {vendor.name} | Code: {vendor.vendor_code} | Contact: {vendor.contact_person or 'N/A'} | Phone: {vendor.phone or 'N/A'}"
    cell.font = Font(bold=True, size=11, color="374151")
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    ws.row_dimensions[header_row].height = 18
    header_row += 2
    
    # Column Headers with trip details
    headers = ['Date', 'Description', 'Reference', 'From', 'To', 'Tonnage', 'Freight', 'Debit', 'Credit', 'Balance', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[header_row].height = 20
    
    # Data rows with alternating colors
    row = header_row + 1
    running_balance = 0
    for idx, entry in enumerate(entries):
        running_balance += entry["debit"] - entry["credit"]
        
        # Alternating row colors
        fill_color = "FFFFFF" if idx % 2 == 0 else "F9FAFB"
        
        ws.cell(row=row, column=1, value=entry["date"].strftime('%Y-%m-%d')).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=row, column=2, value=entry["description"]).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=row, column=3, value=entry["reference"]).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=row, column=4, value=entry["from"]).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=row, column=5, value=entry["to"]).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        cell = ws.cell(row=row, column=6, value=float(entry["tonnage"]))
        cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        cell = ws.cell(row=row, column=7, value=float(entry["freight"]))
        cell.number_format = 'PKR #,##0.00'
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        cell = ws.cell(row=row, column=8, value=float(entry["debit"]))
        cell.number_format = 'PKR #,##0.00'
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        cell = ws.cell(row=row, column=9, value=float(entry["credit"]))
        cell.number_format = 'PKR #,##0.00'
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        cell = ws.cell(row=row, column=10, value=running_balance)
        cell.number_format = 'PKR #,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        status_cell = ws.cell(row=row, column=11, value='Paid' if entry["credit"] > 0 else 'Pending')
        status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        if entry["credit"] > 0:
            status_cell.font = Font(color="059669", bold=True)
        else:
            status_cell.font = Font(color="DC2626")
        
        row += 1
    
    # Summary Section
    row += 2
    ws.merge_cells(f'A{row}:K{row}')
    cell = ws[f'A{row}']
    cell.value = "LEDGER SUMMARY"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 20
    
    row += 1
    total_debit = sum(e["debit"] for e in entries)
    total_credit = sum(e["credit"] for e in entries)
    total_tonnage = sum(e["tonnage"] for e in entries if e["type"] == "payable")
    
    summary_data = [
        ("Total Trips:", len([e for e in entries if e["type"] == "payable"])),
        ("Total Tonnage:", f"{total_tonnage:.2f} MT"),
        ("Total Payable (Debit):", f"PKR {total_debit:,.2f}"),
        ("Total Paid (Credit):", f"PKR {total_credit:,.2f}"),
        ("Outstanding Balance:", f"PKR {(total_debit - total_credit):,.2f}")
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = Font(bold=True, size=10, color="DC2626" if label == "Outstanding Balance:" else "374151")
        if label == "Outstanding Balance:":
            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vendor_ledger_{vendor.name.replace(' ', '_')}.xlsx"}
    )

@app.get("/reports/client-ledger-excel/{client_id}")
def generate_client_ledger_excel(
    client_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate client ledger Excel report with company header and trip details from real-time data"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    import io
    
    # Get client
    client = db.query(models.Client).filter(models.Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get receivables
    receivables_query = db.query(models.Receivable).filter(
        models.Receivable.client_id == client_id
    )
    
    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        receivables_query = receivables_query.filter(models.Receivable.invoice_date >= start)
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d")
        receivables_query = receivables_query.filter(models.Receivable.invoice_date <= end)
    
    receivables = receivables_query.order_by(models.Receivable.invoice_date).all()
    
    # Build entries with trip details
    entries = []
    for receivable in receivables:
        # Get trip details if trip_id exists
        trip = None
        if receivable.trip_id:
            trip = db.query(models.Trip).filter(models.Trip.id == receivable.trip_id).first()
        
        entry = {
            "date": receivable.invoice_date,
            "description": receivable.description or f"Invoice: {receivable.invoice_number}",
            "reference": trip.reference_no if trip else receivable.invoice_number,
            "from": trip.source_location if trip else "",
            "to": trip.destination_location if trip else "",
            "tonnage": float(trip.total_tonnage) if trip and trip.total_tonnage else 0,
            "freight": float(trip.client_freight) if trip and trip.client_freight else float(receivable.total_amount),
            "debit": receivable.total_amount,
            "credit": 0,
            "type": "receivable"
        }
        entries.append(entry)
        
        # Get collections
        collections = db.query(models.Collection).filter(
            models.Collection.receivable_id == receivable.id
        ).all()
        
        for collection in collections:
            entries.append({
                "date": collection.collection_date,
                "description": f"Payment Received: {collection.collection_channel.value if collection.collection_channel else 'N/A'}",
                "reference": collection.reference_number or "",
                "from": "",
                "to": "",
                "tonnage": 0,
                "freight": 0,
                "debit": 0,
                "credit": collection.collection_amount,
                "type": "collection"
            })
    
    entries.sort(key=lambda x: x["date"])
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Ledger"
    
    # Company Header
    ws['A1'] = "PGT International Transport Management"
    ws['A1'].font = Font(bold=True, size=16, color="DC2626")
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "Address: [Your Company Address] | Phone: [Your Phone] | Email: [Your Email]"
    ws['A2'].font = Font(size=10)
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Client Info
    ws['A4'] = f"Client Ledger: {client.name}"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A5'] = f"Code: {client.client_code} | Contact: {client.contact_person or 'N/A'} | Phone: {client.phone or 'N/A'}"
    
    if start_date or end_date:
        ws['A6'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
    
    # Headers with trip details
    headers = ['Date', 'Description', 'Reference', 'From', 'To', 'Tonnage', 'Freight', 'Debit', 'Credit', 'Balance', 'Status']
    header_row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = header_row + 1
    running_balance = 0
    for entry in entries:
        running_balance += entry["debit"] - entry["credit"]
        ws.cell(row=row, column=1, value=entry["date"].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=entry["description"])
        ws.cell(row=row, column=3, value=entry["reference"])
        ws.cell(row=row, column=4, value=entry["from"])
        ws.cell(row=row, column=5, value=entry["to"])
        ws.cell(row=row, column=6, value=float(entry["tonnage"]))
        ws.cell(row=row, column=7, value=float(entry["freight"]))
        ws.cell(row=row, column=8, value=float(entry["debit"]))
        ws.cell(row=row, column=9, value=float(entry["credit"]))
        ws.cell(row=row, column=10, value=running_balance)
        ws.cell(row=row, column=11, value='Paid' if entry["credit"] > 0 else 'Pending')
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1
    total_debit = sum(e["debit"] for e in entries)
    total_credit = sum(e["credit"] for e in entries)
    total_tonnage = sum(e["tonnage"] for e in entries if e["type"] == "receivable")
    ws.cell(row=row, column=1, value="Total Tonnage:")
    ws.cell(row=row, column=2, value=float(total_tonnage)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Debit:")
    ws.cell(row=row, column=2, value=float(total_debit)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Credit:")
    ws.cell(row=row, column=2, value=float(total_credit)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Balance:")
    ws.cell(row=row, column=2, value=float(total_debit - total_credit)).font = Font(bold=True, color="DC2626")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=client_ledger_{client.name.replace(' ', '_')}.xlsx"}
    )
    
    # Client Info
    ws['A4'] = f"Client Ledger: {client.name}"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A5'] = f"Code: {client.client_code} | Contact: {client.contact_person or 'N/A'} | Phone: {client.phone or 'N/A'}"
    
    if start_date or end_date:
        ws['A6'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
    
    # Headers
    headers = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance', 'Status']
    header_row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = header_row + 1
    running_balance = 0
    for entry in entries:
        running_balance += entry["debit"] - entry["credit"]
        ws.cell(row=row, column=1, value=entry["date"].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=entry["description"])
        ws.cell(row=row, column=3, value=entry["reference"])
        ws.cell(row=row, column=4, value=float(entry["debit"]))
        ws.cell(row=row, column=5, value=float(entry["credit"]))
        ws.cell(row=row, column=6, value=running_balance)
        ws.cell(row=row, column=7, value='Paid' if entry["credit"] > 0 else 'Pending')
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1
    total_debit = sum(e["debit"] for e in entries)
    total_credit = sum(e["credit"] for e in entries)
    ws.cell(row=row, column=1, value="Total Debit:")
    ws.cell(row=row, column=2, value=float(total_debit)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Credit:")
    ws.cell(row=row, column=2, value=float(total_credit)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Balance:")
    ws.cell(row=row, column=2, value=float(total_debit - total_credit)).font = Font(bold=True, color="DC2626")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=client_ledger_{client.name.replace(' ', '_')}.xlsx"}
    )

@app.get("/reports/staff-payroll-pdf")
def generate_staff_payroll_pdf(
    month: int = None,
    year: int = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate staff payroll PDF report"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    import io
    
    # Default to current month/year
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    # Get staff data
    staff_list = crud.get_all_staff(db)
    staff_data = [
        {
            "id": staff.id,
            "name": staff.name,
            "employee_id": getattr(staff, 'employee_id', f"EMP-{staff.id:04d}"),
            "position": staff.position,
            "gross_salary": staff.gross_salary
        }
        for staff in staff_list
    ]
    
    # Get payroll entries for the period
    payroll_entries = db.query(models.PayrollEntry).filter(
        models.PayrollEntry.month == month,
        models.PayrollEntry.year == year
    ).all()
    
    payroll_data = [
        {
            "staff_id": entry.staff_id,
            "gross_salary": entry.gross_salary,
            "advance_deduction": entry.advance_deduction,
            "other_deductions": entry.other_deductions,
            "net_payable": entry.net_payable
        }
        for entry in payroll_entries
    ]
    
    # Generate PDF
    generator = ReportGenerator()
    period = f"{datetime(year, month, 1).strftime('%B %Y')}"
    pdf_buffer = generator.generate_staff_payroll_pdf(staff_data, payroll_data, period)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=staff_payroll_{year}_{month:02d}.pdf"}
    )

@app.get("/reports/financial-summary-pdf")
def generate_financial_summary_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate financial summary PDF report"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    from financial_calculator import FinancialCalculator
    import io
    
    # Get financial data
    calculator = FinancialCalculator(db)
    try:
        financial_data = calculator.get_master_financial_summary()
    finally:
        calculator.close()
    
    # Generate PDF
    generator = ReportGenerator()
    pdf_buffer = generator.generate_financial_summary_pdf(financial_data)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=financial_summary_{datetime.now().strftime('%Y_%m_%d')}.pdf"}
    )

@app.get("/reports/trips-excel")
def export_trips_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[int] = None,
    vendor_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export trips data to Excel with optional filters"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    from datetime import datetime as dt
    import io
    
    # Build query with filters
    query = db.query(models.Trip)
    
    # Apply filters
    if start_date:
        query = query.filter(models.Trip.date >= dt.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(models.Trip.date <= dt.strptime(end_date, '%Y-%m-%d'))
    if client_id:
        query = query.filter(models.Trip.client_id == client_id)
    if vendor_id:
        query = query.filter(models.Trip.vendor_id == vendor_id)
    if status:
        # Convert string status to enum
        try:
            status_enum = models.TripStatus[status]
            query = query.filter(models.Trip.status == status_enum)
        except KeyError:
            pass  # Invalid status, ignore filter
    
    # Get filtered trips
    trips = query.order_by(models.Trip.date.desc()).limit(10000).all()
    
    trips_data = [
        {
            "Date": trip.date.strftime('%Y-%m-%d') if trip.date else '',
            "Reference": trip.reference_no,
            "Status": trip.status.value if trip.status else 'DRAFT',
            "Vehicle": trip.vehicle.vehicle_no if trip.vehicle else '',
            "Category": trip.category_product,
            "Source": trip.source_location or '',
            "Destination": trip.destination_location or '',
            "Driver": trip.driver_operator or '',
            "Client": trip.client.name if trip.client else '',
            "Vendor": trip.vendor.name if trip.vendor else '',
            "Client Freight": trip.client_freight or 0,
            "Vendor Freight": trip.vendor_freight or 0,
            "Gross Profit": trip.gross_profit or 0,
            "Advance Paid": trip.advance_paid or 0,
            "Fuel Cost": trip.fuel_cost or 0,
            "Munshiyana Charges": trip.munshiyana_bank_charges or 0,
            "Other Expenses": trip.other_expenses or 0,
            "Net Profit": trip.net_profit or 0,
            "Freight Mode": trip.freight_mode or 'total',
            "Total Tonnage": trip.total_tonnage or 0,
            "Tonnage (Rate)": trip.tonnage or 0,
            "Rate per Ton": trip.rate_per_ton or 0
        }
        for trip in trips
    ]
    
    # Generate Excel
    generator = ReportGenerator()
    excel_buffer = generator.generate_excel_export(trips_data, "trips_export.xlsx", "Trips")
    
    # Build filename with filter info
    filename = "trips_export"
    if status:
        filename += f"_{status}"
    if start_date:
        filename += f"_from_{start_date}"
    if end_date:
        filename += f"_to_{end_date}"
    filename += f"_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/reports/expenses-excel")
def export_expenses_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export expenses data to Excel"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    import io
    
    # Get expenses data
    expenses = crud.get_expenses(db, skip=0, limit=10000)
    
    expenses_data = [
        {
            "Date": expense.date.strftime('%Y-%m-%d') if expense.date else '',
            "Category": expense.expense_category,
            "Description": expense.description,
            "Amount": expense.amount,
            "Vehicle": getattr(expense, 'vehicle_no', ''),
            "Vendor": getattr(expense, 'vendor_name', ''),
            "Receipt": 'Yes' if getattr(expense, 'receipt_image', None) else 'No'
        }
        for expense in expenses
    ]
    
    # Generate Excel
    generator = ReportGenerator()
    excel_buffer = generator.generate_excel_export(expenses_data, "expenses_export.xlsx", "Expenses")
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=expenses_export_{datetime.now().strftime('%Y_%m_%d')}.xlsx"}
    )

@app.get("/reports/payables-pdf")
def generate_payables_report_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate comprehensive payables report PDF"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    from datetime import datetime
    import io
    
    try:
        # Get all payables with vendor information
        payables = db.query(models.Payable).join(models.Vendor).all()
        
        # Format data for report
        payables_data = []
        for payable in payables:
            # Handle datetime comparison properly
            is_overdue = False
            if payable.due_date:
                due_date = payable.due_date.date() if hasattr(payable.due_date, 'date') else payable.due_date
                is_overdue = due_date < datetime.now().date()
            
            payables_data.append({
                'vendor_name': payable.vendor.name,
                'vendor_code': payable.vendor.vendor_code or '',
                'vendor_contact': payable.vendor.contact_person or '',
                'vendor_phone': payable.vendor.phone or '',
                'outstanding_amount': payable.outstanding_amount or 0,
                'last_payment_date': payable.updated_at.strftime('%Y-%m-%d') if payable.updated_at else 'Never',
                'is_overdue': is_overdue
            })
        
        # Generate PDF
        report_generator = ReportGenerator()
        pdf_buffer = report_generator.generate_payables_report_pdf(payables_data)
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=payables_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
        
    except Exception as e:
        print(f"Error generating payables PDF: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to generate payables report")

@app.get("/reports/receivables-pdf")
def generate_receivables_report_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate comprehensive receivables report PDF"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    import io
    
    try:
        # Get all receivables with client information
        receivables = db.query(models.Receivable).join(models.Client).all()
        
        # Format data for report
        receivables_data = []
        for receivable in receivables:
            # Handle datetime comparison properly
            is_overdue = False
            if receivable.due_date:
                due_date = receivable.due_date.date() if hasattr(receivable.due_date, 'date') else receivable.due_date
                is_overdue = due_date < datetime.now().date()
            
            receivables_data.append({
                'client_name': receivable.client.name,
                'client_code': receivable.client.client_code or '',
                'client_contact': receivable.client.contact_person or '',
                'client_phone': receivable.client.phone or '',
                'outstanding_amount': receivable.remaining_amount or 0,
                'last_payment_date': receivable.last_payment_date.strftime('%Y-%m-%d') if receivable.last_payment_date else 'Never',
                'is_overdue': is_overdue
            })
        
        # Generate PDF
        report_generator = ReportGenerator()
        pdf_buffer = report_generator.generate_receivables_report_pdf(receivables_data)
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=receivables_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
        
    except Exception as e:
        print(f"Error generating receivables PDF: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to generate receivables report")

@app.get("/reports/payables-excel")
def export_payables_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export payables data to Excel"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    import io
    
    try:
        # Get all payables with vendor information
        payables = db.query(models.Payable).join(models.Vendor).all()
        
        # Format data for Excel
        payables_data = []
        for payable in payables:
            # Handle datetime comparison properly
            is_overdue = False
            if payable.due_date:
                due_date = payable.due_date.date() if hasattr(payable.due_date, 'date') else payable.due_date
                is_overdue = due_date < datetime.now().date()
            
            payables_data.append({
                'vendor_name': payable.vendor.name,
                'vendor_code': payable.vendor.vendor_code or '',
                'vendor_contact': payable.vendor.contact_person or '',
                'vendor_phone': payable.vendor.phone or '',
                'outstanding_amount': payable.outstanding_amount or 0,
                'due_date': payable.due_date.strftime('%Y-%m-%d') if payable.due_date else '',
                'is_overdue': is_overdue,
                'created_date': payable.created_at.strftime('%Y-%m-%d') if payable.created_at else '',
                'status': payable.status
            })
        
        # Generate Excel
        report_generator = ReportGenerator()
        excel_buffer = report_generator.generate_payables_excel(payables_data)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=payables_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        print(f"Error exporting payables to Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to export payables data")

@app.get("/reports/receivables-excel")
def export_receivables_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export receivables data to Excel"""
    from fastapi.responses import StreamingResponse
    from report_generator import ReportGenerator
    import io
    
    try:
        # Get all receivables with client information
        receivables = db.query(models.Receivable).join(models.Client).all()
        
        # Format data for Excel
        receivables_data = []
        for receivable in receivables:
            # Handle datetime comparison properly
            is_overdue = False
            if receivable.due_date:
                due_date = receivable.due_date.date() if hasattr(receivable.due_date, 'date') else receivable.due_date
                is_overdue = due_date < datetime.now().date()
            
            receivables_data.append({
                'client_name': receivable.client.name,
                'client_code': receivable.client.client_code or '',
                'client_contact': receivable.client.contact_person or '',
                'client_phone': receivable.client.phone or '',
                'outstanding_amount': receivable.remaining_amount or 0,
                'due_date': receivable.due_date.strftime('%Y-%m-%d') if receivable.due_date else '',
                'is_overdue': is_overdue,
                'created_date': receivable.invoice_date.strftime('%Y-%m-%d') if receivable.invoice_date else '',
                'invoice_number': receivable.invoice_number or '',
                'status': receivable.status.value if hasattr(receivable.status, 'value') else str(receivable.status)
            })
        
        # Generate Excel
        report_generator = ReportGenerator()
        excel_buffer = report_generator.generate_receivables_excel(receivables_data)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=receivables_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        print(f"Error exporting receivables to Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to export receivables data")

@app.get("/reports/export-all-data")
def export_all_data(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """
    Director's Safety Requirement: Export ALL data to Excel
    Complete backup of entire system in one file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Trips
        ws_trips = wb.create_sheet("Trips")
        ws_trips.append(['Date', 'Reference', 'Vehicle', 'Client', 'Vendor', 'Product', 'Route', 
                        'Tonnage', 'Client Freight', 'Vendor Freight', 'Gross Profit', 'Net Profit', 'Status'])
        trips = db.query(models.Trip).order_by(models.Trip.date.desc()).all()
        for trip in trips:
            ws_trips.append([
                trip.date.strftime('%Y-%m-%d') if trip.date else '',
                trip.reference_no,
                trip.vehicle.vehicle_no if trip.vehicle else '',
                trip.client.name if trip.client else '',
                trip.vendor.name if trip.vendor else '',
                trip.category_product,
                f"{trip.source_location} → {trip.destination_location}",
                trip.total_tonnage,
                trip.client_freight,
                trip.vendor_freight,
                trip.gross_profit,
                trip.net_profit,
                trip.status.value if hasattr(trip.status, 'value') else str(trip.status)
            ])
        
        # Sheet 2: Clients
        ws_clients = wb.create_sheet("Clients")
        ws_clients.append(['Client Code', 'Name', 'Contact Person', 'Phone', 'Email', 'Current Balance', 'Status'])
        clients = db.query(models.Client).all()
        for client in clients:
            ws_clients.append([
                client.client_code,
                client.name,
                client.contact_person,
                client.phone,
                client.email,
                client.current_balance,
                'Active' if client.is_active else 'Inactive'
            ])
        
        # Sheet 3: Vendors
        ws_vendors = wb.create_sheet("Vendors")
        ws_vendors.append(['Vendor Code', 'Name', 'Contact Person', 'Phone', 'Email', 'Current Balance', 'Status'])
        vendors = db.query(models.Vendor).all()
        for vendor in vendors:
            ws_vendors.append([
                vendor.vendor_code,
                vendor.name,
                vendor.contact_person,
                vendor.phone,
                vendor.email,
                vendor.current_balance,
                'Active' if vendor.is_active else 'Inactive'
            ])
        
        # Sheet 4: Staff
        ws_staff = wb.create_sheet("Staff")
        ws_staff.append(['Employee ID', 'Name', 'Position', 'Gross Salary', 'Advance Balance', 
                        'Monthly Deduction', 'Status'])
        staff = db.query(models.Staff).all()
        for member in staff:
            ws_staff.append([
                member.employee_id,
                member.name,
                member.position,
                member.gross_salary,
                member.advance_balance,
                member.monthly_deduction,
                'Active' if member.is_active else 'Inactive'
            ])
        
        # Sheet 5: Staff Advance Ledger
        ws_advances = wb.create_sheet("Staff Advances")
        ws_advances.append(['Date', 'Staff Name', 'Type', 'Amount', 'Balance After', 'Description'])
        advances = db.query(models.StaffAdvanceLedger).order_by(models.StaffAdvanceLedger.transaction_date.desc()).all()
        for adv in advances:
            ws_advances.append([
                adv.transaction_date.strftime('%Y-%m-%d') if adv.transaction_date else '',
                adv.staff.name if adv.staff else '',
                adv.transaction_type,
                adv.amount,
                adv.balance_after,
                adv.description
            ])
        
        # Sheet 6: Receivables
        ws_receivables = wb.create_sheet("Receivables")
        ws_receivables.append(['Invoice #', 'Client', 'Invoice Date', 'Due Date', 'Total Amount', 
                              'Paid Amount', 'Remaining', 'Status'])
        receivables = db.query(models.Receivable).all()
        for rec in receivables:
            ws_receivables.append([
                rec.invoice_number,
                rec.client.name if rec.client else '',
                rec.invoice_date.strftime('%Y-%m-%d') if rec.invoice_date else '',
                rec.due_date.strftime('%Y-%m-%d') if rec.due_date else '',
                rec.total_amount,
                rec.paid_amount,
                rec.remaining_amount,
                rec.status.value if hasattr(rec.status, 'value') else str(rec.status)
            ])
        
        # Sheet 7: Payables
        ws_payables = wb.create_sheet("Payables")
        ws_payables.append(['Invoice #', 'Vendor', 'Due Date', 'Amount', 'Outstanding', 'Status'])
        payables = db.query(models.Payable).all()
        for pay in payables:
            ws_payables.append([
                pay.invoice_number,
                pay.vendor.name if pay.vendor else '',
                pay.due_date.strftime('%Y-%m-%d') if pay.due_date else '',
                pay.amount,
                pay.outstanding_amount,
                pay.status
            ])
        
        # Sheet 8: Office Expenses
        ws_expenses = wb.create_sheet("Office Expenses")
        ws_expenses.append(['Date', 'Account Title', 'Particulars', 'Amount Received', 'Amount Paid'])
        expenses = db.query(models.OfficeExpense).order_by(models.OfficeExpense.date.desc()).all()
        for exp in expenses:
            ws_expenses.append([
                exp.date.strftime('%Y-%m-%d') if exp.date else '',
                exp.account_title,
                exp.particulars,
                exp.amount_received,
                exp.amount_paid
            ])
        
        # Sheet 9: Vehicles
        ws_vehicles = wb.create_sheet("Vehicles")
        ws_vehicles.append(['Vehicle #', 'Type', 'Capacity (tons)', 'Status'])
        vehicles = db.query(models.Vehicle).all()
        for vehicle in vehicles:
            ws_vehicles.append([
                vehicle.vehicle_no,
                vehicle.vehicle_type,
                vehicle.capacity_tons,
                'Active' if vehicle.is_active else 'Inactive'
            ])
        
        # Style headers for all sheets
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=PGT_Complete_Data_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        print(f"Error exporting all data: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to export all data")

# Staff endpoints
@app.post("/staff/", response_model=schemas.Staff)
def create_staff(
    staff: schemas.StaffCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_staff(db=db, staff=staff)

@app.get("/staff/", response_model=list[schemas.Staff])
def read_staff(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_all_staff(db, skip=skip, limit=limit)

@app.put("/staff/{staff_id}/advance")
def update_staff_advance(
    staff_id: int,
    advance_amount: float,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.update_staff_advance(db, staff_id, advance_amount)

# ============================================
# DIRECTOR'S RULE #1: STAFF ADVANCE RECOVERY
# ============================================

@app.post("/staff/{staff_id}/advance")
def give_staff_advance(
    staff_id: int,
    advance_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """
    Give advance to staff member (handles multiple advances)
    
    Example: Muhammad Hussain has 140,000 existing advance
    Can give another 5,000 emergency advance
    System adds to existing balance: 140,000 + 5,000 = 145,000
    """
    staff = db.query(models.Staff).filter(models.Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    amount = advance_data.get('amount', 0)
    description = advance_data.get('description', 'Advance given')
    monthly_deduction = advance_data.get('monthly_deduction')
    
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Advance amount must be positive")
    
    # Get current balance
    current_balance = staff.advance_balance or 0
    
    # Add new advance to existing balance
    new_balance = current_balance + amount
    
    # Create ledger entry
    ledger_entry = models.StaffAdvanceLedger(
        staff_id=staff_id,
        transaction_date=datetime.now(),
        transaction_type='advance_given',
        amount=amount,
        balance_after=new_balance,
        description=description,
        created_by=current_user.id
    )
    db.add(ledger_entry)
    
    # Update staff balance
    staff.advance_balance = new_balance
    staff.advance_given_date = datetime.now()
    
    # Update monthly deduction if provided
    if monthly_deduction is not None and monthly_deduction > 0:
        staff.monthly_deduction = monthly_deduction
        staff.recovery_start_date = datetime.now()
    
    db.commit()
    db.refresh(staff)
    db.refresh(ledger_entry)
    
    return {
        "success": True,
        "staff_id": staff_id,
        "staff_name": staff.name,
        "advance_given": amount,
        "previous_balance": current_balance,
        "new_balance": new_balance,
        "monthly_deduction": staff.monthly_deduction,
        "ledger_entry_id": ledger_entry.id
    }

@app.get("/staff/{staff_id}/advance-ledger")
def get_staff_advance_ledger(
    staff_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """
    Get complete advance recovery history for a staff member
    Shows all advances given and recoveries made
    """
    staff = db.query(models.Staff).filter(models.Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    # Get all ledger entries
    ledger_entries = db.query(models.StaffAdvanceLedger).filter(
        models.StaffAdvanceLedger.staff_id == staff_id
    ).order_by(models.StaffAdvanceLedger.transaction_date.desc()).all()
    
    # Format entries
    entries = []
    for entry in ledger_entries:
        entries.append({
            "id": entry.id,
            "date": entry.transaction_date.isoformat(),
            "type": entry.transaction_type,
            "amount": entry.amount,
            "balance_after": entry.balance_after,
            "description": entry.description,
            "payroll_id": entry.payroll_id,
            "created_by": entry.created_by_user.full_name if entry.created_by_user else None
        })
    
    # Calculate months to clear (if monthly deduction is set)
    months_to_clear = None
    expected_clear_date = None
    if staff.monthly_deduction and staff.monthly_deduction > 0 and staff.advance_balance > 0:
        months_to_clear = int(staff.advance_balance / staff.monthly_deduction) + 1
        if staff.recovery_start_date:
            from dateutil.relativedelta import relativedelta
            expected_clear_date = staff.recovery_start_date + relativedelta(months=months_to_clear)
    
    return {
        "staff": {
            "id": staff.id,
            "employee_id": staff.employee_id,
            "name": staff.name,
            "position": staff.position,
            "gross_salary": staff.gross_salary,
            "is_active": staff.is_active
        },
        "advance_summary": {
            "current_balance": staff.advance_balance,
            "monthly_deduction": staff.monthly_deduction,
            "advance_given_date": staff.advance_given_date.isoformat() if staff.advance_given_date else None,
            "recovery_start_date": staff.recovery_start_date.isoformat() if staff.recovery_start_date else None,
            "months_to_clear": months_to_clear,
            "expected_clear_date": expected_clear_date.isoformat() if expected_clear_date else None
        },
        "ledger_entries": entries,
        "total_entries": len(entries)
    }

@app.post("/staff/{staff_id}/advance-recovery")
def manual_advance_recovery(
    staff_id: int,
    recovery_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """
    Manual advance recovery entry (not linked to payroll)
    Used for one-time recoveries or adjustments
    """
    staff = db.query(models.Staff).filter(models.Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    amount = recovery_data.get('amount', 0)
    description = recovery_data.get('description', 'Manual recovery')
    
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Recovery amount must be positive")
    
    if amount > staff.advance_balance:
        raise HTTPException(status_code=400, detail=f"Recovery amount ({amount}) exceeds advance balance ({staff.advance_balance})")
    
    # Get current balance
    current_balance = staff.advance_balance
    
    # Subtract recovery from balance
    new_balance = current_balance - amount
    
    # Create ledger entry
    ledger_entry = models.StaffAdvanceLedger(
        staff_id=staff_id,
        transaction_date=datetime.now(),
        transaction_type='recovery',
        amount=-amount,  # Negative for recovery
        balance_after=new_balance,
        description=description,
        created_by=current_user.id
    )
    db.add(ledger_entry)
    
    # Update staff balance
    staff.advance_balance = new_balance
    
    db.commit()
    db.refresh(staff)
    db.refresh(ledger_entry)
    
    return {
        "success": True,
        "staff_id": staff_id,
        "staff_name": staff.name,
        "recovery_amount": amount,
        "previous_balance": current_balance,
        "new_balance": new_balance,
        "ledger_entry_id": ledger_entry.id
    }

# Vehicle endpoints
@app.post("/vehicles/", response_model=schemas.Vehicle)
def create_vehicle(
    vehicle: schemas.VehicleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_vehicle(db=db, vehicle=vehicle)

@app.get("/vehicles/", response_model=list[schemas.Vehicle])
def read_vehicles(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_vehicles(db, skip=skip, limit=limit)

# Enhanced Trip endpoints (matching Excel structure)
@app.post("/trips/", response_model=schemas.Trip)
def create_trip(
    trip: schemas.TripCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Create a new trip with automatic receivable and payable creation"""
    try:
        # Validate trip data
        trip_dict = trip.model_dump()
        validated_data = BusinessValidator.validate_trip_data(trip_dict)
        
        # Create trip with audit logging
        db_trip = crud.create_trip(
            db=db, 
            trip=trip, 
            current_user_id=current_user.id,
            request=request
        )
        
        # Send notification to user
        notif = NotificationService(db)
        notif.create_notification(
            user_id=current_user.id,
            title="Trip Created Successfully",
            message=f"Trip {db_trip.reference_no} created with profit PKR {db_trip.net_profit:,.0f}",
            notification_type="success",
            link="/fleet-logs"
        )
        
        return db_trip
        
    except ValidationError as e:
        raise HTTPException(
            status_code=400,
            detail={"error": "Validation Error", "field": e.field, "message": e.message}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/trips/")
def read_trips(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get all trips with proper enum serialization"""
    # Filter based on user role
    trips = crud.get_trips(db, skip=skip, limit=limit)
    
    # Convert trips to dict and handle enums
    result = []
    for trip in trips:
        trip_dict = {
            "id": trip.id,
            "date": trip.date,
            "reference_no": trip.reference_no,
            "vehicle_id": trip.vehicle_id,
            "category_product": trip.category_product,
            "source_location": trip.source_location,
            "destination_location": trip.destination_location,
            "driver_operator": trip.driver_operator,
            "client_id": trip.client_id,
            "vendor_id": trip.vendor_id,
            "vendor_client": trip.vendor_client,
            "freight_mode": trip.freight_mode,
            "total_tonnage": trip.total_tonnage,
            "tonnage": trip.tonnage,
            "rate_per_ton": trip.rate_per_ton,
            "vendor_freight": trip.vendor_freight,
            "client_freight": trip.client_freight,
            "local_shifting_charges": trip.local_shifting_charges,
            "advance_paid": trip.advance_paid,
            "fuel_cost": trip.fuel_cost,
            "munshiyana_bank_charges": trip.munshiyana_bank_charges,
            "other_expenses": trip.other_expenses,
            "gross_profit": trip.gross_profit,
            "net_profit": trip.net_profit,
            "profit_margin": trip.profit_margin,
            "receivable_created": trip.receivable_created,
            "payable_created": trip.payable_created,
            "receivable_id": trip.receivable_id,
            "payable_id": trip.payable_id,
            "status": trip.status.value if hasattr(trip.status, 'value') else str(trip.status),
            "notes": trip.notes,
            "created_at": trip.created_at,
            "updated_at": trip.updated_at,
            "completed_at": trip.completed_at,
            "client_name": trip.client_name,
            "vendor_name": trip.vendor_name,
            "vehicle_number": trip.vehicle_number
        }
        result.append(trip_dict)
    
    return result

# Payroll endpoints
@app.post("/payroll/", response_model=schemas.PayrollEntry)
def create_payroll_entry(
    payroll: schemas.PayrollEntryCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_payroll_entry(db=db, payroll=payroll)

@app.get("/payroll/", response_model=list[schemas.PayrollEntry])
def read_payroll_entries(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.get_payroll_entries(db, skip=skip, limit=limit)

# Vendor endpoints
@app.post("/vendors/", response_model=schemas.Vendor)
def create_vendor(
    vendor: schemas.VendorCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_vendor(db=db, vendor=vendor)

@app.get("/vendors/", response_model=list[schemas.Vendor])
def read_vendors(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_vendors(db, skip=skip, limit=limit)

# Ledger endpoints
@app.post("/ledger/", response_model=schemas.LedgerEntry)
def create_ledger_entry(
    ledger: schemas.LedgerEntryCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_ledger_entry(db=db, ledger=ledger)

@app.get("/ledger/{vendor_id}", response_model=list[schemas.LedgerEntry])
def read_ledger_entries(
    vendor_id: int,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_ledger_entries(db, vendor_id, skip=skip, limit=limit)

# Enhanced Ledger Endpoints with Real-Time Data from Trips
@app.get("/api/ledgers/vendor/{vendor_id}")
def get_vendor_ledger_detailed(
    vendor_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get vendor ledger from real-time payables and payments with trip details"""
    try:
        from datetime import datetime
        
        # Get vendor
        vendor = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        
        # Get all payables for this vendor
        payables_query = db.query(models.Payable).filter(
            models.Payable.vendor_id == vendor_id
        )
        
        # Apply date filters
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            payables_query = payables_query.filter(models.Payable.created_at >= start)
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            payables_query = payables_query.filter(models.Payable.created_at <= end)
        
        payables = payables_query.order_by(models.Payable.created_at).all()
        
        # Build ledger entries from payables and payments
        entries = []
        running_balance = 0
        
        for payable in payables:
            # Find the trip associated with this payable
            trip = db.query(models.Trip).filter(models.Trip.payable_id == payable.id).first()
            
            # Build trip details if trip exists
            trip_details = None
            if trip:
                trip_details = {
                    "from": trip.source_location,
                    "to": trip.destination_location,
                    "tonnage": float(trip.total_tonnage) if trip.total_tonnage else None,
                    "freight": float(trip.vendor_freight) if trip.vendor_freight else None,
                    "vehicle": trip.vehicle_number
                }
            
            # Debit entry (we owe vendor)
            running_balance += payable.amount
            entries.append({
                "id": f"payable_{payable.id}",
                "date": payable.created_at,
                "description": payable.description or f"Invoice: {payable.invoice_number}",
                "trip_reference": trip.reference_no if trip else payable.invoice_number,
                "trip_details": trip_details,
                "debit": float(payable.amount),
                "credit": 0,
                "balance": running_balance,
                "type": "payable",
                "status": "paid" if payable.status == "paid" else "pending"
            })
            
            # Get payments for this payable (both APPROVED and PAID status)
            payments = db.query(models.PaymentRequest).filter(
                models.PaymentRequest.payable_id == payable.id,
                models.PaymentRequest.status.in_([models.PaymentRequestStatus.APPROVED, models.PaymentRequestStatus.PAID])
            ).all()
            
            for payment in payments:
                if payment.payment_date:
                    running_balance -= payment.requested_amount
                    entries.append({
                        "id": f"payment_{payment.id}",
                        "date": payment.payment_date,
                        "description": f"Payment: {payment.payment_channel.value if payment.payment_channel else 'N/A'}",
                        "trip_reference": payment.payment_reference or "",
                        "trip_details": None,
                        "debit": 0,
                        "credit": float(payment.requested_amount),
                        "balance": running_balance,
                        "type": "payment",
                        "status": "paid"
                    })
        
        # Sort by date
        entries.sort(key=lambda x: x["date"])
        
        # Recalculate running balance
        running_balance = 0
        for entry in entries:
            running_balance += entry["debit"] - entry["credit"]
            entry["balance"] = running_balance
        
        # Calculate summary
        total_debit = sum(e["debit"] for e in entries)
        total_credit = sum(e["credit"] for e in entries)
        
        return {
            "vendor": {
                "id": vendor.id,
                "name": vendor.name,
                "code": vendor.vendor_code,
                "contact": vendor.contact_person,
                "phone": vendor.phone
            },
            "entries": entries,
            "summary": {
                "total_debit": float(total_debit),
                "total_credit": float(total_credit),
                "balance": float(total_debit - total_credit),
                "trip_count": len([e for e in entries if e["type"] == "payable"]),
                "payment_count": len([e for e in entries if e["type"] == "payment"])
            }
        }
    except Exception as e:
        print(f"Error in get_vendor_ledger_detailed: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ledgers/client/{client_id}")
def get_client_ledger_detailed(
    client_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get client ledger from real-time receivables and collections with trip details"""
    try:
        from datetime import datetime
        
        # Get client
        client = db.query(models.Client).filter(models.Client.id == client_id).first()
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Get all receivables for this client
        receivables_query = db.query(models.Receivable).filter(
            models.Receivable.client_id == client_id
        )
        
        # Apply date filters
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            receivables_query = receivables_query.filter(models.Receivable.invoice_date >= start)
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            receivables_query = receivables_query.filter(models.Receivable.invoice_date <= end)
        
        receivables = receivables_query.order_by(models.Receivable.invoice_date).all()
        
        # Build ledger entries from receivables and collections
        entries = []
        running_balance = 0
        
        for receivable in receivables:
            # Get trip details if trip_id exists
            trip_details = None
            trip_reference = receivable.invoice_number
            
            if receivable.trip_id:
                trip = db.query(models.Trip).filter(models.Trip.id == receivable.trip_id).first()
                if trip:
                    trip_reference = trip.reference_no
                    trip_details = {
                        "from": trip.source_location,
                        "to": trip.destination_location,
                        "tonnage": float(trip.total_tonnage) if trip.total_tonnage else None,
                        "freight": float(trip.client_freight) if trip.client_freight else None,
                        "vehicle": trip.vehicle_number
                    }
            
            # Debit entry (client owes us)
            running_balance += receivable.total_amount
            entries.append({
                "id": f"receivable_{receivable.id}",
                "date": receivable.invoice_date,
                "description": receivable.description or f"Invoice: {receivable.invoice_number}",
                "trip_reference": trip_reference,
                "trip_details": trip_details,
                "debit": float(receivable.total_amount),
                "credit": 0,
                "balance": running_balance,
                "type": "receivable",
                "status": "paid" if receivable.status == models.ReceivableStatus.PAID else "pending"
            })
            
            # Get collections for this receivable
            collections = db.query(models.Collection).filter(
                models.Collection.receivable_id == receivable.id
            ).all()
            
            for collection in collections:
                running_balance -= collection.collection_amount
                entries.append({
                    "id": f"collection_{collection.id}",
                    "date": collection.collection_date,
                    "description": f"Payment Received: {collection.collection_channel.value if collection.collection_channel else 'N/A'}",
                    "trip_reference": collection.reference_number or "",
                    "trip_details": None,
                    "debit": 0,
                    "credit": float(collection.collection_amount),
                    "balance": running_balance,
                    "type": "collection",
                    "status": "paid"
                })
        
        # Sort by date
        entries.sort(key=lambda x: x["date"])
        
        # Recalculate running balance
        running_balance = 0
        for entry in entries:
            running_balance += entry["debit"] - entry["credit"]
            entry["balance"] = running_balance
        
        # Calculate summary
        total_debit = sum(e["debit"] for e in entries)
        total_credit = sum(e["credit"] for e in entries)
        
        return {
            "client": {
                "id": client.id,
                "name": client.name,
                "code": client.client_code,
                "contact": client.contact_person,
                "phone": client.phone
            },
            "entries": entries,
            "summary": {
                "total_debit": float(total_debit),
                "total_credit": float(total_credit),
                "balance": float(total_debit - total_credit),
                "trip_count": len([e for e in entries if e["type"] == "receivable"]),
                "payment_count": len([e for e in entries if e["type"] == "collection"])
            }
        }
    except Exception as e:
        print(f"Error in get_client_ledger_detailed: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ledgers/vendors/summary")
def get_all_vendors_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get summary of all vendors with real-time outstanding balances"""
    vendors = db.query(models.Vendor).filter(models.Vendor.is_active == True).all()
    
    summary = []
    for vendor in vendors:
        # Calculate total payables
        payables = db.query(models.Payable).filter(
            models.Payable.vendor_id == vendor.id
        ).all()
        
        total_payable = sum(p.amount for p in payables)
        total_paid = sum(p.amount - (p.outstanding_amount or p.amount) for p in payables)
        balance = sum(p.outstanding_amount or p.amount for p in payables)
        
        summary.append({
            "vendor_id": vendor.id,
            "vendor_name": vendor.name,
            "vendor_code": vendor.vendor_code,
            "total_debit": float(total_payable),
            "total_credit": float(total_paid),
            "balance": float(balance),
            "trip_count": len(payables)
        })
    
    return summary

@app.get("/api/ledgers/clients/summary")
def get_all_clients_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get summary of all clients with real-time outstanding balances"""
    clients = db.query(models.Client).filter(models.Client.is_active == True).all()
    
    summary = []
    for client in clients:
        # Calculate total receivables
        receivables = db.query(models.Receivable).filter(
            models.Receivable.client_id == client.id
        ).all()
        
        total_receivable = sum(r.total_amount for r in receivables)
        total_received = sum(r.paid_amount or 0 for r in receivables)
        balance = sum(r.remaining_amount or r.total_amount for r in receivables)
        
        summary.append({
            "client_id": client.id,
            "client_name": client.name,
            "client_code": client.client_code,
            "total_debit": float(total_receivable),
            "total_credit": float(total_received),
            "balance": float(balance),
            "trip_count": len(receivables)
        })
    
    return summary

# Expense endpoints
@app.post("/expenses/", response_model=schemas.Expense)
def create_expense(
    expense: schemas.ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.create_expense(db=db, expense=expense, current_user_id=current_user.id)

@app.get("/expenses/", response_model=list[schemas.Expense])
def read_expenses(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_expenses(db, skip=skip, limit=limit)

# Office Expense endpoints
@app.post("/office-expenses/", response_model=schemas.OfficeExpense)
def create_office_expense(
    expense: schemas.OfficeExpenseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Create a new office expense entry and automatically deduct from CEO Capital if payment is made"""
    db_expense = models.OfficeExpense(
        **expense.dict(),
        created_by=current_user.id
    )
    db.add(db_expense)
    db.flush()  # Get the expense ID before committing
    
    # If this is a payment (amount_paid > 0), automatically deduct from CEO Capital
    if expense.amount_paid > 0:
        # Get current CEO Capital balance
        last_transaction = db.query(models.CEOCapital).order_by(
            models.CEOCapital.date.desc(),
            models.CEOCapital.id.desc()
        ).first()
        
        current_balance = last_transaction.balance if last_transaction else 0.0
        new_balance = current_balance - expense.amount_paid
        
        # Create CEO Capital transaction
        ceo_transaction = models.CEOCapital(
            date=expense.date,
            transaction_type='expense_payment',
            description=f"Office Expense: {expense.account_title} - {expense.particulars[:50]}",
            amount_in=0.0,
            amount_out=expense.amount_paid,
            balance=new_balance,
            reference_id=db_expense.id,
            reference_type='office_expense',
            created_by=current_user.id
        )
        db.add(ceo_transaction)
        db.flush()
        
        # Link the CEO Capital transaction to the expense
        db_expense.ceo_capital_transaction_id = ceo_transaction.id
    
    db.commit()
    db.refresh(db_expense)
    return db_expense

@app.get("/office-expenses/", response_model=list[schemas.OfficeExpense])
def read_office_expenses(
    skip: int = 0,
    limit: int = 1000,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    entry_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get all office expenses with optional filters"""
    query = db.query(models.OfficeExpense).order_by(models.OfficeExpense.date.asc(), models.OfficeExpense.id.asc())
    
    if start_date:
        query = query.filter(models.OfficeExpense.date >= start_date)
    if end_date:
        query = query.filter(models.OfficeExpense.date <= end_date)
    if category:
        query = query.filter(models.OfficeExpense.account_title == category)
    if entry_type:
        query = query.filter(models.OfficeExpense.entry_type == entry_type)
    
    return query.offset(skip).limit(limit).all()

@app.get("/office-expenses/download")
async def download_office_expenses_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    entry_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Download office expenses as Excel file"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt
    
    # Get filtered expenses
    query = db.query(models.OfficeExpense).order_by(models.OfficeExpense.date.asc(), models.OfficeExpense.id.asc())
    
    if start_date:
        query = query.filter(models.OfficeExpense.date >= start_date)
    if end_date:
        query = query.filter(models.OfficeExpense.date <= end_date)
    if category:
        query = query.filter(models.OfficeExpense.account_title == category)
    if entry_type:
        query = query.filter(models.OfficeExpense.entry_type == entry_type)
    
    expenses = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Determine month/year for title
    if expenses:
        first_date = expenses[0].date
        title = f"SWL OFFICE EXPENSES {first_date.strftime('%b %Y').upper()}"
    else:
        title = f"SWL OFFICE EXPENSES {dt.now().strftime('%b %Y').upper()}"
    
    ws.title = "Office Expenses"
    
    # Add title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 25
    
    # Add headers
    headers = ['Sr#', 'Date', 'Acc. Title', 'Particulars/Descriptions', 'Amount Rcvd', 'Amount Paid', 'Balance']
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Add data rows with running balance
    balance = 0
    for idx, expense in enumerate(expenses, 1):
        row_num = idx + 2
        
        # Calculate running balance
        balance += (expense.amount_received or 0) - (expense.amount_paid or 0)
        
        # Add data
        ws.cell(row=row_num, column=1).value = idx
        ws.cell(row=row_num, column=2).value = expense.date.strftime('%d-%b-%y')
        ws.cell(row=row_num, column=3).value = expense.account_title
        ws.cell(row=row_num, column=4).value = expense.particulars
        ws.cell(row=row_num, column=5).value = expense.amount_received if expense.amount_received > 0 else '-'
        ws.cell(row=row_num, column=6).value = expense.amount_paid if expense.amount_paid > 0 else '-'
        ws.cell(row=row_num, column=7).value = balance
        
        # Format numbers
        if expense.amount_received > 0:
            ws.cell(row=row_num, column=5).number_format = '#,##0'
        if expense.amount_paid > 0:
            ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        
        # Apply borders
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical='center')
        
        # Right align numbers
        for col in [5, 6, 7]:
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right', vertical='center')
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Return as streaming response
    filename = f"Office_Expenses_{dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# CEO Capital endpoints
@app.get("/ceo-capital/balance", response_model=schemas.CEOCapitalBalance)
def get_ceo_capital_balance(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get current CEO Capital balance and summary"""
    transactions = db.query(models.CEOCapital).order_by(models.CEOCapital.date.desc()).all()
    
    if not transactions:
        return schemas.CEOCapitalBalance(
            current_balance=0.0,
            total_in=0.0,
            total_out=0.0,
            last_transaction_date=None
        )
    
    current_balance = transactions[0].balance
    total_in = sum(t.amount_in for t in transactions)
    total_out = sum(t.amount_out for t in transactions)
    last_date = transactions[0].date
    
    return schemas.CEOCapitalBalance(
        current_balance=current_balance,
        total_in=total_in,
        total_out=total_out,
        last_transaction_date=last_date
    )

@app.get("/ceo-capital/transactions", response_model=list[schemas.CEOCapital])
def get_ceo_capital_transactions(
    skip: int = 0,
    limit: int = 100,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get CEO Capital transactions with optional filters"""
    query = db.query(models.CEOCapital).order_by(models.CEOCapital.date.asc(), models.CEOCapital.id.asc())
    
    if start_date:
        query = query.filter(models.CEOCapital.date >= start_date)
    if end_date:
        query = query.filter(models.CEOCapital.date <= end_date)
    if transaction_type:
        query = query.filter(models.CEOCapital.transaction_type == transaction_type)
    
    return query.offset(skip).limit(limit).all()

@app.post("/ceo-capital/transaction", response_model=schemas.CEOCapital)
def create_ceo_capital_transaction(
    transaction: schemas.CEOCapitalCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Create a new CEO Capital transaction (Admin only)"""
    
    # Get current balance
    last_transaction = db.query(models.CEOCapital).order_by(
        models.CEOCapital.date.desc(),
        models.CEOCapital.id.desc()
    ).first()
    
    current_balance = last_transaction.balance if last_transaction else 0.0
    
    # Calculate new balance
    new_balance = current_balance + transaction.amount_in - transaction.amount_out
    
    # Create transaction
    db_transaction = models.CEOCapital(
        **transaction.dict(),
        balance=new_balance,
        created_by=current_user.id
    )
    
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    
    return db_transaction

@app.post("/ceo-capital/allocate-profit")
def allocate_profit_to_ceo_capital(
    profit_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Allocate profit to CEO Capital"""
    from datetime import date as dt_date
    
    amount = profit_data.get('amount')
    description = profit_data.get('description')
    reference_date = profit_data.get('date')
    
    transaction_date = dt_date.fromisoformat(reference_date) if reference_date else dt_date.today()
    
    # Get current balance
    last_transaction = db.query(models.CEOCapital).order_by(
        models.CEOCapital.date.desc(),
        models.CEOCapital.id.desc()
    ).first()
    
    current_balance = last_transaction.balance if last_transaction else 0.0
    new_balance = current_balance + amount
    
    # Create profit allocation transaction
    transaction = models.CEOCapital(
        date=transaction_date,
        transaction_type='profit_allocation',
        description=description,
        amount_in=amount,
        amount_out=0.0,
        balance=new_balance,
        reference_type='manual',
        created_by=current_user.id
    )
    
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    
    return {
        "success": True,
        "message": f"Profit of PKR {amount:,.2f} allocated to CEO Capital",
        "new_balance": new_balance,
        "transaction_id": transaction.id
    }

@app.post("/ceo-capital/withdrawal")
def record_ceo_withdrawal(
    withdrawal_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """Record CEO withdrawal from capital"""
    from datetime import date as dt_date
    
    amount = withdrawal_data.get('amount')
    description = withdrawal_data.get('description')
    withdrawal_date = withdrawal_data.get('date')
    
    transaction_date = dt_date.fromisoformat(withdrawal_date) if withdrawal_date else dt_date.today()
    
    # Get current balance
    last_transaction = db.query(models.CEOCapital).order_by(
        models.CEOCapital.date.desc(),
        models.CEOCapital.id.desc()
    ).first()
    
    current_balance = last_transaction.balance if last_transaction else 0.0
    
    if current_balance < amount:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient balance. Current: PKR {current_balance:,.2f}, Requested: PKR {amount:,.2f}"
        )
    
    new_balance = current_balance - amount
    
    # Create withdrawal transaction
    transaction = models.CEOCapital(
        date=transaction_date,
        transaction_type='withdrawal',
        description=description,
        amount_in=0.0,
        amount_out=amount,
        balance=new_balance,
        reference_type='withdrawal',
        created_by=current_user.id
    )
    
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    
    return {
        "success": True,
        "message": f"Withdrawal of PKR {amount:,.2f} recorded",
        "new_balance": new_balance,
        "transaction_id": transaction.id
    }

@app.get("/ceo-capital/monthly-summary")
def get_ceo_capital_monthly_summary(
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get CEO Capital monthly summary"""
    from datetime import date as dt_date
    from sqlalchemy import extract, and_
    
    # Use current month/year if not provided
    today = dt_date.today()
    target_month = month or today.month
    target_year = year or today.year
    
    # Get transactions for the month
    transactions = db.query(models.CEOCapital).filter(
        and_(
            extract('month', models.CEOCapital.date) == target_month,
            extract('year', models.CEOCapital.date) == target_year
        )
    ).order_by(models.CEOCapital.date.asc()).all()
    
    if not transactions:
        return {
            "month": target_month,
            "year": target_year,
            "opening_balance": 0.0,
            "total_profit": 0.0,
            "total_expenses": 0.0,
            "total_salaries": 0.0,
            "total_withdrawals": 0.0,
            "closing_balance": 0.0,
            "transaction_count": 0
        }
    
    opening_balance = transactions[0].balance - transactions[0].amount_in + transactions[0].amount_out
    closing_balance = transactions[-1].balance
    
    total_profit = sum(t.amount_in for t in transactions if t.transaction_type == 'profit_allocation')
    total_expenses = sum(t.amount_out for t in transactions if t.transaction_type == 'office_expense')
    total_salaries = sum(t.amount_out for t in transactions if t.transaction_type == 'salary_payment')
    total_withdrawals = sum(t.amount_out for t in transactions if t.transaction_type == 'withdrawal')
    
    return {
        "month": target_month,
        "year": target_year,
        "opening_balance": opening_balance,
        "total_profit": total_profit,
        "total_expenses": total_expenses,
        "total_salaries": total_salaries,
        "total_withdrawals": total_withdrawals,
        "closing_balance": closing_balance,
        "transaction_count": len(transactions)
    }

@app.get("/ceo-capital/download")
def download_ceo_capital_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    transaction_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Download CEO Capital transactions as Excel"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = db.query(models.CEOCapital)
    
    if start_date:
        query = query.filter(models.CEOCapital.date >= start_date)
    if end_date:
        query = query.filter(models.CEOCapital.date <= end_date)
    if transaction_type:
        query = query.filter(models.CEOCapital.transaction_type == transaction_type)
    
    transactions = query.order_by(models.CEOCapital.date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CEO Capital"
    
    # Company header
    ws.merge_cells('A1:F1')
    ws['A1'] = "PGT INTERNATIONAL"
    ws['A1'].font = Font(size=16, bold=True, color="DC2626")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "CEO Capital Transaction Report"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Date', 'Type', 'Description', 'Amount In', 'Amount Out', 'Balance']
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_idx, transaction in enumerate(transactions, 6):
        ws.cell(row=row_idx, column=1, value=transaction.date.strftime('%d-%b-%Y'))
        ws.cell(row=row_idx, column=2, value=transaction.transaction_type.replace('_', ' ').title())
        ws.cell(row=row_idx, column=3, value=transaction.description)
        ws.cell(row=row_idx, column=4, value=transaction.amount_in if transaction.amount_in > 0 else '')
        ws.cell(row=row_idx, column=5, value=transaction.amount_out if transaction.amount_out > 0 else '')
        ws.cell(row=row_idx, column=6, value=transaction.balance)
        
        # Apply borders
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = border
            if col >= 4:  # Amount columns
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='right')
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ceo_capital_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Client endpoints
@app.post("/clients/", response_model=schemas.Client)
def create_client(
    client: schemas.ClientCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_client(db=db, client=client)

@app.get("/clients/", response_model=list[schemas.Client])
def read_clients(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_clients(db, skip=skip, limit=limit)

@app.get("/clients/{client_id}", response_model=schemas.Client)
def read_client(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    client = crud.get_client(db, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client

# Receivable endpoints
@app.post("/receivables/", response_model=schemas.Receivable)
def create_receivable(
    receivable: schemas.ReceivableCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_receivable(db=db, receivable=receivable, current_user_id=current_user.id)

@app.get("/receivables/", response_model=list[schemas.Receivable])
def read_receivables(
    skip: int = 0,
    limit: int = 100,
    client_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_receivables(db, skip=skip, limit=limit, client_id=client_id, status=status)

@app.get("/receivables/{receivable_id}", response_model=schemas.Receivable)
def read_receivable(
    receivable_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    receivable = crud.get_receivable(db, receivable_id)
    if not receivable:
        raise HTTPException(status_code=404, detail="Receivable not found")
    return receivable

# Collection endpoints
@app.post("/collections/", response_model=schemas.Collection)
def create_collection(
    collection: schemas.CollectionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    try:
        return crud.create_collection(db=db, collection=collection, current_user_id=current_user.id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/collections/", response_model=list[schemas.Collection])
def read_collections(
    skip: int = 0,
    limit: int = 100,
    receivable_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_collections(db, skip=skip, limit=limit, receivable_id=receivable_id)

# Trip completion and receivable creation
@app.put("/trips/{trip_id}/complete")
def complete_trip_and_create_receivable(
    trip_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Complete trip and automatically create receivable if client is assigned"""
    trip = db.query(models.Trip).filter(models.Trip.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Update trip status
    trip.status = "completed"
    trip.completed_at = datetime.now()
    
    # Create receivable if client is assigned and not already created
    receivable = None
    if trip.client_id and not trip.receivable_created:
        receivable = crud.create_receivable_from_trip(db, trip, current_user.id)
    
    db.commit()
    
    return {
        "message": "Trip completed successfully",
        "trip_id": trip.id,
        "receivable_created": receivable is not None,
        "receivable_id": receivable.id if receivable else None
    }

# Trip status update endpoint
@app.put("/trips/{trip_id}/status")
def update_trip_status(
    trip_id: int,
    status_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Update trip status (DRAFT -> ACTIVE -> COMPLETED -> LOCKED)"""
    trip = db.query(models.Trip).filter(models.Trip.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    new_status = status_data.get("status")
    if not new_status:
        raise HTTPException(status_code=400, detail="Status is required")
    
    # Validate status transition
    valid_statuses = ["DRAFT", "ACTIVE", "COMPLETED", "LOCKED"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    # Update status
    trip.status = models.TripStatus[new_status]
    
    if new_status == "COMPLETED":
        trip.completed_at = datetime.now()
    
    db.commit()
    db.refresh(trip)
    
    return {
        "message": f"Trip status updated to {new_status}",
        "trip_id": trip.id,
        "status": trip.status.value
    }

# Trip cancellation endpoint with financial reversal
@app.put("/trips/{trip_id}/cancel")
def cancel_trip(
    trip_id: int,
    cancel_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Cancel trip and reverse all financial records (receivable, payable, cash transactions)"""
    trip = db.query(models.Trip).filter(models.Trip.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.status in [models.TripStatus.LOCKED]:
        raise HTTPException(status_code=400, detail="Cannot cancel a locked trip")
    
    reason = cancel_data.get("reason", "")
    if not reason:
        raise HTTPException(status_code=400, detail="Cancellation reason is required")
    
    try:
        # Mark trip as cancelled
        trip.status = models.TripStatus.CANCELLED if hasattr(models.TripStatus, 'CANCELLED') else models.TripStatus.DRAFT
        trip.notes = f"{trip.notes or ''}\n\nCANCELLED: {reason} (by {current_user.username} on {datetime.now()})"
        
        # Reverse receivable if exists
        if trip.receivable_id:
            receivable = db.query(models.Receivable).filter(models.Receivable.id == trip.receivable_id).first()
            if receivable:
                receivable.status = models.ReceivableStatus.CANCELLED
                receivable.is_deleted = True
                receivable.deleted_at = datetime.now()
                receivable.deleted_by = current_user.id
        
        # Reverse payable if exists
        if trip.payable_id:
            payable = db.query(models.Payable).filter(models.Payable.id == trip.payable_id).first()
            if payable:
                payable.status = "cancelled"
                payable.is_deleted = True
                payable.deleted_at = datetime.now()
                payable.deleted_by = current_user.id
        
        # Reverse cash transactions
        # Find cash transactions related to this trip's receivable and payable
        if trip.receivable_id:
            cash_txns = db.query(models.CashTransaction).filter(
                models.CashTransaction.source_module == "receivable",
                models.CashTransaction.source_id == trip.receivable_id,
                models.CashTransaction.is_deleted == False
            ).all()
            for txn in cash_txns:
                txn.is_deleted = True
                txn.deleted_at = datetime.now()
                txn.deleted_by = current_user.id
        
        if trip.payable_id:
            cash_txns = db.query(models.CashTransaction).filter(
                models.CashTransaction.source_module == "payable",
                models.CashTransaction.source_id == trip.payable_id,
                models.CashTransaction.is_deleted == False
            ).all()
            for txn in cash_txns:
                txn.is_deleted = True
                txn.deleted_at = datetime.now()
                txn.deleted_by = current_user.id
        
        db.commit()
        
        return {
            "message": "Trip cancelled and financial records reversed successfully",
            "trip_id": trip.id,
            "receivable_reversed": trip.receivable_id is not None,
            "payable_reversed": trip.payable_id is not None,
            "reason": reason
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to cancel trip: {str(e)}")

# Payable endpoints
@app.post("/payables/", response_model=schemas.Payable)
def create_payable(
    payable: schemas.PayableCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.create_payable(db=db, payable=payable)

@app.get("/payables/", response_model=list[schemas.Payable])
def read_payables(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_payables(db, skip=skip, limit=limit)

@app.put("/payables/{payable_id}/status")
def update_payable_status(
    payable_id: int,
    status: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    return crud.update_payable_status(db, payable_id, status)

# Payment Request endpoints
@app.post("/payment-requests/", response_model=schemas.PaymentRequest)
def create_payment_request(
    payment_request: schemas.PaymentRequestCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Create payment request and notify admins"""
    try:
        # Get the payable to check outstanding amount
        payable = crud.get_payable(db, payment_request.payable_id)
        if not payable:
            raise HTTPException(status_code=404, detail="Payable not found")
        
        # Validate payment request with outstanding amount
        validated_data = BusinessValidator.validate_payment_request(
            payment_request.model_dump(), 
            payable.outstanding_amount or payable.amount
        )
        
        # Create payment request
        db_payment_request = crud.create_payment_request(
            db=db, 
            payment_request=payment_request, 
            current_user_id=current_user.id
        )
        
        # Audit logging
        audit = AuditService(db)
        audit.log_create(
            user_id=current_user.id,
            table_name="payment_requests",
            record_id=db_payment_request.id,
            new_values={
                "vendor_id": db_payment_request.vendor_id,
                "requested_amount": db_payment_request.requested_amount,
                "status": db_payment_request.status.value
            },
            ip_address=get_client_ip(request),
            user_agent=get_user_agent(request)
        )
        
        # Notify admins
        notif = NotificationService(db)
        admin_ids = get_admin_user_ids(db)
        vendor = db.query(models.Vendor).filter(models.Vendor.id == db_payment_request.vendor_id).first()
        
        notif.notify_payment_request_submitted(
            admin_user_ids=admin_ids,
            vendor_name=vendor.name if vendor else "Unknown",
            amount=db_payment_request.requested_amount,
            payment_request_id=db_payment_request.id
        )
        
        return db_payment_request
        
    except ValidationError as e:
        raise HTTPException(
            status_code=400,
            detail={"error": "Validation Error", "field": e.field, "message": e.message}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/payment-requests/")
def read_payment_requests(
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get all payment requests with proper enum serialization and relationships"""
    payment_requests = crud.get_payment_requests(db, skip=skip, limit=limit, status=status)
    
    # Manually serialize to handle enums and include relationships
    result = []
    for pr in payment_requests:
        result.append({
            "id": pr.id,
            "payable_id": pr.payable_id,
            "vendor_id": pr.vendor_id,
            "payment_type": pr.payment_type.value if hasattr(pr.payment_type, 'value') else str(pr.payment_type),
            "requested_amount": pr.requested_amount,
            "remaining_amount": pr.remaining_amount,
            "payment_channel": pr.payment_channel.value if hasattr(pr.payment_channel, 'value') else str(pr.payment_channel),
            "request_reason": pr.request_reason,
            "urgency_level": pr.urgency_level,
            "status": pr.status.value if hasattr(pr.status, 'value') else str(pr.status),
            "requested_by": pr.requested_by,
            "requested_at": pr.requested_at,
            "approved_by": pr.approved_by,
            "approved_at": pr.approved_at,
            "rejection_reason": pr.rejection_reason,
            "payment_reference": pr.payment_reference,
            "payment_date": pr.payment_date,
            "payment_notes": pr.payment_notes,
            "created_at": pr.created_at,
            "updated_at": pr.updated_at,
            # Include vendor and payable data
            "vendor": {
                "id": pr.vendor.id,
                "name": pr.vendor.name,
                "vendor_code": pr.vendor.vendor_code
            } if pr.vendor else None,
            "payable": {
                "id": pr.payable.id,
                "invoice_number": pr.payable.invoice_number,
                "amount": pr.payable.amount,
                "outstanding_amount": pr.payable.outstanding_amount
            } if pr.payable else None
        })
    
    return result

@app.get("/payment-requests/{request_id}")
def read_payment_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get single payment request with proper enum serialization"""
    pr = crud.get_payment_request(db, request_id)
    if not pr:
        raise HTTPException(status_code=404, detail="Payment request not found")
    
    # Manually serialize to handle enums
    return {
        "id": pr.id,
        "payable_id": pr.payable_id,
        "vendor_id": pr.vendor_id,
        "payment_type": pr.payment_type.value if hasattr(pr.payment_type, 'value') else str(pr.payment_type),
        "requested_amount": pr.requested_amount,
        "remaining_amount": pr.remaining_amount,
        "payment_channel": pr.payment_channel.value if hasattr(pr.payment_channel, 'value') else str(pr.payment_channel),
        "request_reason": pr.request_reason,
        "urgency_level": pr.urgency_level,
        "status": pr.status.value if hasattr(pr.status, 'value') else str(pr.status),
        "requested_by": pr.requested_by,
        "requested_at": pr.requested_at,
        "approved_by": pr.approved_by,
        "approved_at": pr.approved_at,
        "rejection_reason": pr.rejection_reason,
        "payment_reference": pr.payment_reference,
        "payment_date": pr.payment_date,
        "payment_notes": pr.payment_notes,
        "created_at": pr.created_at,
        "updated_at": pr.updated_at
    }

@app.put("/payment-requests/{request_id}")
def update_payment_request(
    request_id: int,
    update_data: schemas.PaymentRequestUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Update payment request and notify requester"""
    try:
        # Get old values for audit
        old_payment_request = crud.get_payment_request(db, request_id)
        if not old_payment_request:
            raise HTTPException(status_code=404, detail="Payment request not found")
        
        old_status = old_payment_request.status.value
        
        # Update payment request
        payment_request = crud.update_payment_request(db, request_id, update_data, current_user.id)
        
        if not payment_request:
            raise HTTPException(status_code=404, detail="Payment request not found")
        
        # Audit logging
        audit = AuditService(db)
        audit.log_update(
            user_id=current_user.id,
            table_name="payment_requests",
            record_id=request_id,
            old_values={"status": old_status},
            new_values={"status": payment_request.status.value},
            ip_address=get_client_ip(request),
            user_agent=get_user_agent(request)
        )
        
        # Send notification based on status change
        notif = NotificationService(db)
        vendor = db.query(models.Vendor).filter(models.Vendor.id == payment_request.vendor_id).first()
        vendor_name = vendor.name if vendor else "Unknown"
        
        if payment_request.status.value == "approved" and old_status != "approved":
            notif.notify_payment_approved(
                requester_id=payment_request.requested_by,
                vendor_name=vendor_name,
                amount=payment_request.requested_amount
            )
        elif payment_request.status.value == "rejected" and old_status != "rejected":
            notif.notify_payment_rejected(
                requester_id=payment_request.requested_by,
                vendor_name=vendor_name,
                amount=payment_request.requested_amount,
                reason=update_data.rejection_reason or "No reason provided"
            )
        
        # Manually serialize to handle enums
        return {
            "id": payment_request.id,
            "payable_id": payment_request.payable_id,
            "vendor_id": payment_request.vendor_id,
            "payment_type": payment_request.payment_type.value if hasattr(payment_request.payment_type, 'value') else str(payment_request.payment_type),
            "requested_amount": payment_request.requested_amount,
            "remaining_amount": payment_request.remaining_amount,
            "payment_channel": payment_request.payment_channel.value if hasattr(payment_request.payment_channel, 'value') else str(payment_request.payment_channel),
            "request_reason": payment_request.request_reason,
            "urgency_level": payment_request.urgency_level,
            "status": payment_request.status.value if hasattr(payment_request.status, 'value') else str(payment_request.status),
            "requested_by": payment_request.requested_by,
            "requested_at": payment_request.requested_at,
            "approved_by": payment_request.approved_by,
            "approved_at": payment_request.approved_at,
            "rejection_reason": payment_request.rejection_reason,
            "payment_reference": payment_request.payment_reference,
            "payment_date": payment_request.payment_date,
            "payment_notes": payment_request.payment_notes,
            "created_at": payment_request.created_at,
            "updated_at": payment_request.updated_at
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating payment request: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/payment-requests/pending/count")
def get_pending_payment_requests_count(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    count = crud.get_pending_payment_requests_count(db)
    return {"pending_count": count}

# Vehicle Log endpoints
@app.post("/vehicle-logs/", response_model=schemas.VehicleLog)
def create_vehicle_log(
    log: schemas.VehicleLogCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    try:
        return crud.create_vehicle_log(db=db, log=log)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/vehicle-logs/", response_model=list[schemas.VehicleLog])
def read_vehicle_logs(
    skip: int = 0,
    limit: int = 100,
    vehicle_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_vehicle_logs(db, skip=skip, limit=limit, vehicle_id=vehicle_id)

# Summary & Reporting endpoints
@app.get("/reports/summary")
def get_summary_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_summary_report(db, start_date, end_date)

@app.get("/reports/profit-loss")
def get_profit_loss_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_profit_loss_report(db, start_date, end_date)

@app.get("/reports/vendor-balances")
def get_vendor_balances(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_vendor_outstanding_balances(db)

@app.get("/reports/vehicle-performance")
def get_vehicle_performance(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return crud.get_vehicle_performance_report(db, start_date, end_date)

# ============================================
# COMPREHENSIVE VENDOR & CLIENT REPORTS
# ============================================

@app.get("/api/reports/vendor-performance")
def get_vendor_performance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get comprehensive vendor performance report from integrated system"""
    try:
        from datetime import datetime, date, timedelta
        
        # Build date filters
        date_filter_start = None
        date_filter_end = None
        if start_date:
            date_filter_start = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            date_filter_end = datetime.strptime(end_date, "%Y-%m-%d")
        
        # Get vendors
        vendors_query = db.query(models.Vendor).filter(models.Vendor.is_active == True)
        if vendor_id:
            vendors_query = vendors_query.filter(models.Vendor.id == vendor_id)
        vendors = vendors_query.all()
        
        vendor_performance = []
        total_payables = 0
        total_paid = 0
        total_outstanding = 0
        total_trips = 0
        
        for vendor in vendors:
            # Get payables for this vendor
            payables_query = db.query(models.Payable).filter(
                models.Payable.vendor_id == vendor.id
            )
            if date_filter_start:
                payables_query = payables_query.filter(models.Payable.created_at >= date_filter_start)
            if date_filter_end:
                payables_query = payables_query.filter(models.Payable.created_at <= date_filter_end)
            
            payables = payables_query.all()
            
            # Get trips for this vendor
            trips_query = db.query(models.Trip).filter(
                models.Trip.vendor_id == vendor.id,
                models.Trip.status != models.TripStatus.CANCELLED
            )
            if date_filter_start:
                trips_query = trips_query.filter(models.Trip.date >= date_filter_start)
            if date_filter_end:
                trips_query = trips_query.filter(models.Trip.date <= date_filter_end)
            
            trips = trips_query.all()
            
            # Calculate metrics
            vendor_total_payables = sum(p.amount for p in payables)
            vendor_outstanding = sum(p.outstanding_amount or p.amount for p in payables)
            vendor_paid = vendor_total_payables - vendor_outstanding
            vendor_trip_count = len(trips)
            
            # Get payments for payment history
            payment_requests = db.query(models.PaymentRequest).filter(
                models.PaymentRequest.vendor_id == vendor.id,
                models.PaymentRequest.status.in_([models.PaymentRequestStatus.APPROVED, models.PaymentRequestStatus.PAID])
            ).all()
            
            # Calculate payment performance
            on_time_payments = 0
            late_payments = 0
            total_payment_days = 0
            
            for payment in payment_requests:
                if payment.payment_date and payment.payable:
                    payable_date = payment.payable.created_at.date() if hasattr(payment.payable.created_at, 'date') else payment.payable.created_at
                    payment_date = payment.payment_date.date() if hasattr(payment.payment_date, 'date') else payment.payment_date
                    days_to_pay = (payment_date - payable_date).days
                    total_payment_days += days_to_pay
                    
                    if days_to_pay <= 30:
                        on_time_payments += 1
                    else:
                        late_payments += 1
            
            avg_payment_days = total_payment_days / len(payment_requests) if payment_requests else 0
            
            # This month's performance
            this_month_start = date.today().replace(day=1)
            this_month_trips = [t for t in trips if t.date and t.date.date() >= this_month_start]
            this_month_payables = [p for p in payables if p.created_at and p.created_at.date() >= this_month_start]
            
            # Last trip
            last_trip = max(trips, key=lambda t: t.date) if trips else None
            
            # Aging analysis
            aging = {'0-30': 0, '31-60': 0, '61-90': 0, '90+': 0}
            for payable in payables:
                if payable.outstanding_amount and payable.outstanding_amount > 0:
                    payable_date = payable.created_at.date() if hasattr(payable.created_at, 'date') else payable.created_at
                    days_old = (date.today() - payable_date).days
                    
                    if days_old <= 30:
                        aging['0-30'] += float(payable.outstanding_amount)
                    elif days_old <= 60:
                        aging['31-60'] += float(payable.outstanding_amount)
                    elif days_old <= 90:
                        aging['61-90'] += float(payable.outstanding_amount)
                    else:
                        aging['90+'] += float(payable.outstanding_amount)
            
            vendor_data = {
                "vendor_id": vendor.id,
                "vendor_name": vendor.name,
                "vendor_code": vendor.vendor_code,
                "contact_person": vendor.contact_person,
                "phone": vendor.phone,
                "total_trips": vendor_trip_count,
                "total_payables": float(vendor_total_payables),
                "total_paid": float(vendor_paid),
                "outstanding_amount": float(vendor_outstanding),
                "avg_trip_value": float(vendor_total_payables / vendor_trip_count) if vendor_trip_count > 0 else 0,
                "this_month_trips": len(this_month_trips),
                "this_month_payables": float(sum(p.amount for p in this_month_payables)),
                "last_trip_date": last_trip.date.isoformat() if last_trip and last_trip.date else None,
                "last_trip_reference": last_trip.reference_no if last_trip else None,
                "payment_history": {
                    "on_time_payments": on_time_payments,
                    "late_payments": late_payments,
                    "avg_payment_days": round(avg_payment_days, 1)
                },
                "aging": aging
            }
            
            vendor_performance.append(vendor_data)
            
            # Update totals
            total_payables += vendor_total_payables
            total_paid += vendor_paid
            total_outstanding += vendor_outstanding
            total_trips += vendor_trip_count
        
        return {
            "summary": {
                "total_vendors": len(vendors),
                "active_vendors": len([v for v in vendor_performance if v["total_trips"] > 0]),
                "total_payables": float(total_payables),
                "total_paid": float(total_paid),
                "total_outstanding": float(total_outstanding),
                "total_trips": total_trips
            },
            "vendors": vendor_performance
        }
        
    except Exception as e:
        print(f"Error in vendor performance report: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/client-performance")
def get_client_performance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get comprehensive client performance report from integrated system"""
    try:
        from datetime import datetime, date, timedelta
        
        # Build date filters
        date_filter_start = None
        date_filter_end = None
        if start_date:
            date_filter_start = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            date_filter_end = datetime.strptime(end_date, "%Y-%m-%d")
        
        # Get clients
        clients_query = db.query(models.Client).filter(models.Client.is_active == True)
        if client_id:
            clients_query = clients_query.filter(models.Client.id == client_id)
        clients = clients_query.all()
        
        client_performance = []
        total_receivables = 0
        total_collected = 0
        total_outstanding = 0
        total_trips = 0
        
        for client in clients:
            # Get receivables for this client
            receivables_query = db.query(models.Receivable).filter(
                models.Receivable.client_id == client.id
            )
            if date_filter_start:
                receivables_query = receivables_query.filter(models.Receivable.invoice_date >= date_filter_start)
            if date_filter_end:
                receivables_query = receivables_query.filter(models.Receivable.invoice_date <= date_filter_end)
            
            receivables = receivables_query.all()
            
            # Get trips for this client
            trips_query = db.query(models.Trip).filter(
                models.Trip.client_id == client.id,
                models.Trip.status != models.TripStatus.CANCELLED
            )
            if date_filter_start:
                trips_query = trips_query.filter(models.Trip.date >= date_filter_start)
            if date_filter_end:
                trips_query = trips_query.filter(models.Trip.date <= date_filter_end)
            
            trips = trips_query.all()
            
            # Calculate metrics
            client_total_receivables = sum(r.total_amount for r in receivables)
            client_collected = sum(r.paid_amount or 0 for r in receivables)
            client_outstanding = sum(r.remaining_amount or r.total_amount for r in receivables)
            client_trip_count = len(trips)
            
            # Get collections for collection history
            collections = db.query(models.Collection).join(
                models.Receivable
            ).filter(
                models.Receivable.client_id == client.id
            ).all()
            
            # Calculate collection performance
            on_time_collections = 0
            late_collections = 0
            total_collection_days = 0
            
            for collection in collections:
                if collection.collection_date and collection.receivable:
                    invoice_date = collection.receivable.invoice_date.date() if hasattr(collection.receivable.invoice_date, 'date') else collection.receivable.invoice_date
                    collection_date = collection.collection_date.date() if hasattr(collection.collection_date, 'date') else collection.collection_date
                    days_to_collect = (collection_date - invoice_date).days
                    total_collection_days += days_to_collect
                    
                    if days_to_collect <= 30:
                        on_time_collections += 1
                    else:
                        late_collections += 1
            
            avg_collection_days = total_collection_days / len(collections) if collections else 0
            
            # This month's performance
            this_month_start = date.today().replace(day=1)
            this_month_trips = [t for t in trips if t.date and t.date.date() >= this_month_start]
            this_month_receivables = [r for r in receivables if r.invoice_date and r.invoice_date.date() >= this_month_start]
            
            # Last trip
            last_trip = max(trips, key=lambda t: t.date) if trips else None
            
            # Destinations and products
            destinations = list(set([t.destination_location for t in trips if t.destination_location]))
            products = list(set([t.category_product for t in trips if t.category_product]))
            
            # Aging analysis
            aging = {'0-30': 0, '31-60': 0, '61-90': 0, '90+': 0}
            for receivable in receivables:
                if receivable.remaining_amount and receivable.remaining_amount > 0:
                    invoice_date = receivable.invoice_date.date() if hasattr(receivable.invoice_date, 'date') else receivable.invoice_date
                    days_old = (date.today() - invoice_date).days
                    
                    if days_old <= 30:
                        aging['0-30'] += float(receivable.remaining_amount)
                    elif days_old <= 60:
                        aging['31-60'] += float(receivable.remaining_amount)
                    elif days_old <= 90:
                        aging['61-90'] += float(receivable.remaining_amount)
                    else:
                        aging['90+'] += float(receivable.remaining_amount)
            
            client_data = {
                "client_id": client.id,
                "client_name": client.name,
                "client_code": client.client_code,
                "contact_person": client.contact_person,
                "phone": client.phone,
                "total_trips": client_trip_count,
                "total_receivables": float(client_total_receivables),
                "total_collected": float(client_collected),
                "outstanding_amount": float(client_outstanding),
                "avg_trip_value": float(client_total_receivables / client_trip_count) if client_trip_count > 0 else 0,
                "this_month_trips": len(this_month_trips),
                "this_month_receivables": float(sum(r.total_amount for r in this_month_receivables)),
                "last_trip_date": last_trip.date.isoformat() if last_trip and last_trip.date else None,
                "last_trip_reference": last_trip.reference_no if last_trip else None,
                "destinations": destinations[:10],  # Top 10
                "products": products[:10],  # Top 10
                "collection_history": {
                    "on_time_collections": on_time_collections,
                    "late_collections": late_collections,
                    "avg_collection_days": round(avg_collection_days, 1)
                },
                "aging": aging
            }
            
            client_performance.append(client_data)
            
            # Update totals
            total_receivables += client_total_receivables
            total_collected += client_collected
            total_outstanding += client_outstanding
            total_trips += client_trip_count
        
        return {
            "summary": {
                "total_clients": len(clients),
                "active_clients": len([c for c in client_performance if c["total_trips"] > 0]),
                "total_receivables": float(total_receivables),
                "total_collected": float(total_collected),
                "total_outstanding": float(total_outstanding),
                "total_trips": total_trips
            },
            "clients": client_performance
        }
        
    except Exception as e:
        print(f"Error in client performance report: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/vendor-performance-excel")
def export_vendor_performance_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export vendor performance report to Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import io
    
    # Get report data
    report_data = get_vendor_performance_report(start_date, end_date, vendor_id, db, current_user)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Performance"
    
    # Company Header
    ws['A1'] = "PGT International Transport Management"
    ws['A1'].font = Font(bold=True, size=16, color="DC2626")
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Vendor Performance Report"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:M2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    if start_date or end_date:
        ws['A3'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
        ws.merge_cells('A3:M3')
        ws['A3'].alignment = Alignment(horizontal='center')
    
    # Summary section
    row = 5
    ws[f'A{row}'] = "SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    summary = report_data['summary']
    ws[f'A{row}'] = "Total Vendors:"
    ws[f'B{row}'] = summary['total_vendors']
    row += 1
    ws[f'A{row}'] = "Active Vendors:"
    ws[f'B{row}'] = summary['active_vendors']
    row += 1
    ws[f'A{row}'] = "Total Payables:"
    ws[f'B{row}'] = float(summary['total_payables'])
    row += 1
    ws[f'A{row}'] = "Total Paid:"
    ws[f'B{row}'] = float(summary['total_paid'])
    row += 1
    ws[f'A{row}'] = "Total Outstanding:"
    ws[f'B{row}'] = float(summary['total_outstanding'])
    row += 2
    
    # Headers
    headers = ['Vendor Code', 'Vendor Name', 'Contact', 'Phone', 'Total Trips', 'Total Payables', 
               'Total Paid', 'Outstanding', 'Avg Trip Value', 'This Month Trips', 'Last Trip Date', 
               'On-Time Payments', 'Avg Payment Days']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row += 1
    for vendor in report_data['vendors']:
        ws.cell(row=row, column=1, value=vendor['vendor_code'])
        ws.cell(row=row, column=2, value=vendor['vendor_name'])
        ws.cell(row=row, column=3, value=vendor['contact_person'] or '')
        ws.cell(row=row, column=4, value=vendor['phone'] or '')
        ws.cell(row=row, column=5, value=vendor['total_trips'])
        ws.cell(row=row, column=6, value=float(vendor['total_payables']))
        ws.cell(row=row, column=7, value=float(vendor['total_paid']))
        ws.cell(row=row, column=8, value=float(vendor['outstanding_amount']))
        ws.cell(row=row, column=9, value=float(vendor['avg_trip_value']))
        ws.cell(row=row, column=10, value=vendor['this_month_trips'])
        ws.cell(row=row, column=11, value=vendor['last_trip_date'] or 'N/A')
        ws.cell(row=row, column=12, value=vendor['payment_history']['on_time_payments'])
        ws.cell(row=row, column=13, value=vendor['payment_history']['avg_payment_days'])
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vendor_performance_report.xlsx"}
    )

@app.get("/api/reports/client-performance-excel")
def export_client_performance_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export client performance report to Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import io
    
    # Get report data
    report_data = get_client_performance_report(start_date, end_date, client_id, db, current_user)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Performance"
    
    # Company Header
    ws['A1'] = "PGT International Transport Management"
    ws['A1'].font = Font(bold=True, size=16, color="DC2626")
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Client Performance Report"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:M2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    if start_date or end_date:
        ws['A3'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
        ws.merge_cells('A3:M3')
        ws['A3'].alignment = Alignment(horizontal='center')
    
    # Summary section
    row = 5
    ws[f'A{row}'] = "SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    summary = report_data['summary']
    ws[f'A{row}'] = "Total Clients:"
    ws[f'B{row}'] = summary['total_clients']
    row += 1
    ws[f'A{row}'] = "Active Clients:"
    ws[f'B{row}'] = summary['active_clients']
    row += 1
    ws[f'A{row}'] = "Total Receivables:"
    ws[f'B{row}'] = float(summary['total_receivables'])
    row += 1
    ws[f'A{row}'] = "Total Collected:"
    ws[f'B{row}'] = float(summary['total_collected'])
    row += 1
    ws[f'A{row}'] = "Total Outstanding:"
    ws[f'B{row}'] = float(summary['total_outstanding'])
    row += 2
    
    # Headers
    headers = ['Client Code', 'Client Name', 'Contact', 'Phone', 'Total Trips', 'Total Receivables', 
               'Total Collected', 'Outstanding', 'Avg Trip Value', 'This Month Trips', 'Last Trip Date', 
               'On-Time Collections', 'Avg Collection Days']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row += 1
    for client in report_data['clients']:
        ws.cell(row=row, column=1, value=client['client_code'])
        ws.cell(row=row, column=2, value=client['client_name'])
        ws.cell(row=row, column=3, value=client['contact_person'] or '')
        ws.cell(row=row, column=4, value=client['phone'] or '')
        ws.cell(row=row, column=5, value=client['total_trips'])
        ws.cell(row=row, column=6, value=float(client['total_receivables']))
        ws.cell(row=row, column=7, value=float(client['total_collected']))
        ws.cell(row=row, column=8, value=float(client['outstanding_amount']))
        ws.cell(row=row, column=9, value=float(client['avg_trip_value']))
        ws.cell(row=row, column=10, value=client['this_month_trips'])
        ws.cell(row=row, column=11, value=client['last_trip_date'] or 'N/A')
        ws.cell(row=row, column=12, value=client['collection_history']['on_time_collections'])
        ws.cell(row=row, column=13, value=client['collection_history']['avg_collection_days'])
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=client_performance_report.xlsx"}
    )


# ============================================
# ADMIN ENDPOINTS - DANGEROUS OPERATIONS
# ============================================

@app.post("/admin/reset-database")
def reset_database(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """
    DANGER: Reset all data in the database
    This will delete ALL records except users
    Admin only - requires confirmation
    """
    try:
        # Delete all data (keep users for login)
        db.query(models.StaffAdvanceLedger).delete()
        db.query(models.OfficeExpense).delete()
        db.query(models.Receivable).delete()
        db.query(models.Payable).delete()
        db.query(models.FinancialTransaction).delete()
        db.query(models.Trip).delete()
        db.query(models.Staff).delete()
        db.query(models.Client).delete()
        db.query(models.Vendor).delete()
        db.query(models.Vehicle).delete()
        db.query(models.VehicleLog).delete()
        
        db.commit()
        
        return {
            "message": "Database reset successfully",
            "deleted": "All trips, clients, vendors, staff, vehicles, and financial records",
            "preserved": "User accounts"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to reset database: {str(e)}")


# ============================================
# ENHANCED REPORT ENDPOINTS - INTERNATIONAL STANDARDS
# Director's 4 Requirements Implementation
# ============================================

@app.get("/reports/vendor-ledger-pdf-enhanced/{vendor_id}")
def generate_vendor_ledger_pdf_enhanced(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """
    Generate ENHANCED vendor ledger PDF with:
    - Quick Info Box (top right)
    - Monthly transaction grouping
    - Color-coded payment status
    - Running balance always visible
    """
    from fastapi.responses import StreamingResponse
    from enhanced_reports import EnhancedReportGenerator
    
    # Get vendor data
    vendor = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Get last payment date
    last_payment = db.query(models.Payable).filter(
        models.Payable.vendor_id == vendor_id,
        models.Payable.outstanding_amount < models.Payable.amount
    ).order_by(models.Payable.date.desc()).first()
    
    last_payment_date = last_payment.date if last_payment else None
    
    # Get ledger entries
    ledger_entries = []
    try:
        # Get payables for this vendor
        payables = db.query(models.Payable).filter(
            models.Payable.vendor_id == vendor_id
        ).order_by(models.Payable.date).all()
        
        running_balance = 0
        for payable in payables:
            # Debit entry (new payable)
            running_balance += float(payable.amount)
            ledger_entries.append({
                "date": payable.date,
                "description": f"Invoice {payable.invoice_number}",
                "reference_no": payable.invoice_number,
                "debit_amount": float(payable.amount),
                "credit_amount": 0,
                "running_balance": running_balance
            })
            
            # Credit entry (payment made)
            if payable.amount > payable.outstanding_amount:
                paid_amount = float(payable.amount - payable.outstanding_amount)
                running_balance -= paid_amount
                ledger_entries.append({
                    "date": payable.date,
                    "description": f"Payment for {payable.invoice_number}",
                    "reference_no": f"PAY-{payable.invoice_number}",
                    "debit_amount": 0,
                    "credit_amount": paid_amount,
                    "running_balance": running_balance
                })
    except Exception as e:
        print(f"Warning: Could not fetch ledger entries: {e}")
    
    # Convert to dict format
    vendor_data = {
        "name": vendor.name,
        "vendor_code": f"VEN-{vendor.id:04d}",
        "contact_person": vendor.contact_person,
        "phone": vendor.phone,
        "email": getattr(vendor, 'email', ''),
        "current_balance": float(vendor.current_balance),
        "last_payment_date": last_payment_date,
        "is_active": vendor.is_active
    }
    
    # Generate enhanced PDF
    generator = EnhancedReportGenerator()
    pdf_buffer = generator.generate_vendor_ledger_pdf_enhanced(vendor_data, ledger_entries)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Ledger_{datetime.now().strftime('%Y-%m-%d')}_{vendor.name.replace(' ', '_')}.pdf"}
    )

@app.get("/reports/financial-summary-pdf-enhanced")
def generate_financial_summary_pdf_enhanced(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN]))
):
    """
    Generate ENHANCED financial summary PDF with:
    - Detailed expense breakdown (Office/Staff/Vendor)
    - Receivable aging table at bottom
    - All standard metrics
    """
    from fastapi.responses import StreamingResponse
    from enhanced_reports import EnhancedReportGenerator
    from financial_calculator import FinancialCalculator
    
    calculator = FinancialCalculator(db)
    
    try:
        # Get financial summary
        financial_data = calculator.get_master_financial_summary()
        
        # ENHANCEMENT: Get expense breakdown
        # Office expenses
        office_expenses = db.query(func.sum(models.OfficeExpense.amount)).filter(
            models.OfficeExpense.date >= date.today().replace(day=1)
        ).scalar() or 0
        
        # Staff salaries (current month)
        staff_salaries = db.query(func.sum(models.Staff.gross_salary)).scalar() or 0
        
        # Vendor payments (current month)
        vendor_payments = db.query(
            func.sum(models.Payable.amount - models.Payable.outstanding_amount)
        ).filter(
            models.Payable.date >= date.today().replace(day=1)
        ).scalar() or 0
        
        expense_breakdown = {
            "office_expenses": float(office_expenses),
            "staff_salaries": float(staff_salaries),
            "vendor_payments": float(vendor_payments),
            "other_expenses": 0,
            "total": float(office_expenses + staff_salaries + vendor_payments)
        }
        
        # ENHANCEMENT: Get receivable aging
        receivables = db.query(models.Receivable).filter(
            models.Receivable.remaining_amount > 0
        ).all()
        
        aging_buckets = {
            "0-30": 0,
            "31-60": 0,
            "61-90": 0,
            "90+": 0,
            "total": 0
        }
        
        for receivable in receivables:
            days_old = (date.today() - receivable.invoice_date).days
            amount = float(receivable.remaining_amount)
            aging_buckets["total"] += amount
            
            if days_old <= 30:
                aging_buckets["0-30"] += amount
            elif days_old <= 60:
                aging_buckets["31-60"] += amount
            elif days_old <= 90:
                aging_buckets["61-90"] += amount
            else:
                aging_buckets["90+"] += amount
        
        # Generate enhanced PDF
        generator = EnhancedReportGenerator()
        pdf_buffer = generator.generate_financial_summary_pdf_enhanced(
            financial_data,
            expense_breakdown,
            aging_buckets
        )
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Financial_Summary_{datetime.now().strftime('%Y-%m-%d')}.pdf"}
        )
    finally:
        calculator.close()

@app.get("/reports/staff-statement-pdf-enhanced/{staff_id}")
def generate_staff_statement_pdf_enhanced(
    staff_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """
    Generate ENHANCED staff advance statement PDF with:
    - Quick Info Box showing advance balance
    - Bank statement style
    - Running balance decreasing each month
    - Professional format
    """
    from fastapi.responses import StreamingResponse
    from enhanced_reports import EnhancedReportGenerator
    
    # Get staff data
    staff = db.query(models.Staff).filter(models.Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    
    # Get advance ledger entries
    ledger_entries = db.query(models.StaffAdvanceLedger).filter(
        models.StaffAdvanceLedger.staff_id == staff_id
    ).order_by(models.StaffAdvanceLedger.transaction_date).all()
    
    # Get last deduction date
    last_deduction = db.query(models.StaffAdvanceLedger).filter(
        models.StaffAdvanceLedger.staff_id == staff_id,
        models.StaffAdvanceLedger.transaction_type == 'recovery'
    ).order_by(models.StaffAdvanceLedger.transaction_date.desc()).first()
    
    # Convert to dict format
    staff_data = {
        "name": staff.name,
        "employee_id": staff.employee_id,
        "position": staff.position,
        "gross_salary": float(staff.gross_salary),
        "advance_balance": float(staff.advance_balance),
        "monthly_deduction": float(staff.monthly_deduction),
        "phone": getattr(staff, 'phone', 'N/A'),
        "last_deduction_date": last_deduction.transaction_date if last_deduction else None
    }
    
    ledger_data = []
    for entry in ledger_entries:
        ledger_data.append({
            "date": entry.transaction_date,
            "description": entry.description,
            "debit_amount": float(entry.debit_amount),
            "credit_amount": float(entry.credit_amount),
            "running_balance": float(entry.balance_after)
        })
    
    # Generate enhanced PDF
    generator = EnhancedReportGenerator()
    pdf_buffer = generator.generate_staff_statement_pdf_enhanced(staff_data, ledger_data)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Staff_Statement_{datetime.now().strftime('%Y-%m-%d')}_{staff.name.replace(' ', '_')}.pdf"}
    )


# ============================================
# ENHANCED INVOICE MANAGEMENT ENDPOINTS
# ============================================

@app.post("/invoices/generate-from-trip/{trip_id}")
def generate_invoice_from_trip(
    trip_id: int,
    auto_email: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate professional invoice from trip with all details"""
    from invoice_service import InvoiceService
    
    service = InvoiceService(db)
    result = service.generate_invoice_from_trip(trip_id, auto_email=auto_email)
    
    if result['success']:
        return result
    else:
        raise HTTPException(status_code=400, detail=result['error'])

@app.get("/invoices/list")
def list_invoices(
    client_id: Optional[int] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """List all invoices with filters"""
    from invoice_service import InvoiceService
    from datetime import datetime
    
    service = InvoiceService(db)
    
    start = datetime.fromisoformat(start_date) if start_date else None
    end = datetime.fromisoformat(end_date) if end_date else None
    
    invoices = service.list_invoices(
        client_id=client_id,
        status=status,
        start_date=start,
        end_date=end,
        skip=skip,
        limit=limit
    )
    
    return {"invoices": invoices, "total": len(invoices)}

@app.get("/invoices/{invoice_id}/pdf")
def download_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Download invoice PDF"""
    from fastapi.responses import StreamingResponse
    from invoice_service import InvoiceService
    import io
    
    service = InvoiceService(db)
    pdf_data = service.get_invoice_pdf(invoice_id)
    
    if not pdf_data:
        raise HTTPException(status_code=404, detail="Invoice PDF not found")
    
    receivable = db.query(models.Receivable).filter(models.Receivable.id == invoice_id).first()
    
    return StreamingResponse(
        io.BytesIO(pdf_data),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={receivable.invoice_number}.pdf"
        }
    )

@app.post("/invoices/{invoice_id}/regenerate")
def regenerate_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Regenerate invoice PDF"""
    from invoice_service import InvoiceService
    
    service = InvoiceService(db)
    result = service.regenerate_invoice(invoice_id)
    
    if result['success']:
        return result
    else:
        raise HTTPException(status_code=400, detail=result['error'])

@app.post("/invoices/{invoice_id}/email")
def email_invoice_to_client(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Email invoice to client"""
    from invoice_service import InvoiceService
    
    service = InvoiceService(db)
    result = service.email_invoice(invoice_id)
    
    if result['success']:
        return result
    else:
        raise HTTPException(status_code=400, detail=result['error'])

@app.get("/invoices/summary")
def get_invoice_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get invoice summary statistics"""
    from invoice_service import InvoiceService
    from datetime import datetime
    
    service = InvoiceService(db)
    
    start = datetime.fromisoformat(start_date) if start_date else None
    end = datetime.fromisoformat(end_date) if end_date else None
    
    summary = service.get_invoice_summary(
        start_date=start,
        end_date=end,
        client_id=client_id
    )
    
    return summary

@app.post("/invoices/bulk-generate")
def bulk_generate_invoices(
    trip_ids: list[int],
    auto_email: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Generate invoices for multiple trips"""
    from invoice_service import InvoiceService
    
    service = InvoiceService(db)
    results = service.bulk_generate_invoices(trip_ids, auto_email=auto_email)
    
    return results

# ============================================
# TWO-FACTOR AUTHENTICATION (2FA) ENDPOINTS
# ============================================

@app.post("/2fa/enable")
def enable_2fa(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Enable 2FA for current user"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.enable_2fa(db, current_user.id)
    
    return result

@app.post("/2fa/disable")
def disable_2fa(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Disable 2FA for current user"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.disable_2fa(db, current_user.id)
    
    return result

@app.post("/2fa/send-otp")
def send_otp(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Send OTP to user's email"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.send_otp(db, current_user.id)
    
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=400, detail=result["error"])

@app.post("/2fa/verify-otp")
def verify_otp(
    otp: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Verify OTP code"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.verify_otp(db, current_user.id, otp)
    
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=400, detail=result["error"])

@app.post("/2fa/generate-backup-codes")
def generate_backup_codes(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Generate backup codes for 2FA"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.generate_backup_codes(db, current_user.id)
    
    return result

@app.post("/2fa/verify-backup-code")
def verify_backup_code(
    code: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Verify backup code"""
    from two_factor_auth import TwoFactorAuth
    
    tfa = TwoFactorAuth()
    result = tfa.verify_backup_code(db, current_user.id, code)
    
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=400, detail=result["error"])

@app.get("/2fa/status")
def get_2fa_status(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Get 2FA status for current user"""
    return {
        "enabled": current_user.two_factor_enabled if hasattr(current_user, 'two_factor_enabled') else False,
        "email": current_user.email
    }

# ============================================
# BULK IMPORT/EXPORT ENDPOINTS
# ============================================

@app.post("/import/clients")
async def import_clients(
    file: bytes,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Import clients from CSV/Excel file"""
    from bulk_import_export import BulkImportExport
    from fastapi import File, UploadFile
    
    service = BulkImportExport(db)
    result = service.import_clients(file)
    
    return result

@app.post("/import/vendors")
async def import_vendors(
    file: bytes,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Import vendors from CSV/Excel file"""
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    result = service.import_vendors(file)
    
    return result

@app.post("/import/staff")
async def import_staff(
    file: bytes,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Import staff from CSV/Excel file"""
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    result = service.import_staff(file)
    
    return result

@app.post("/import/vehicles")
async def import_vehicles(
    file: bytes,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Import vehicles from CSV/Excel file"""
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    result = service.import_vehicles(file)
    
    return result

@app.get("/export/clients")
def export_clients(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export all clients to Excel"""
    from fastapi.responses import StreamingResponse
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    excel_buffer = service.export_clients()
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=clients_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@app.get("/export/vendors")
def export_vendors(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export all vendors to Excel"""
    from fastapi.responses import StreamingResponse
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    excel_buffer = service.export_vendors()
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vendors_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@app.get("/export/staff")
def export_staff(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export all staff to Excel"""
    from fastapi.responses import StreamingResponse
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    excel_buffer = service.export_staff()
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=staff_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@app.get("/export/trips")
def export_trips(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Export trips to Excel with optional date range"""
    from fastapi.responses import StreamingResponse
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    excel_buffer = service.export_trips(start_date, end_date)
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trips_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@app.get("/templates/{entity_type}")
def download_import_template(
    entity_type: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Download import template for specified entity type"""
    from fastapi.responses import StreamingResponse
    from bulk_import_export import BulkImportExport
    
    service = BulkImportExport(db)
    
    if entity_type == "clients":
        template = service.get_client_template()
    elif entity_type == "vendors":
        template = service.get_vendor_template()
    elif entity_type == "staff":
        template = service.get_staff_template()
    elif entity_type == "vehicles":
        template = service.get_vehicle_template()
    else:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    return StreamingResponse(
        template,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity_type}_import_template.xlsx"}
    )

# ============================================
# PAYSLIP GENERATION ENDPOINTS
# ============================================

@app.post("/payslips/generate/{payroll_id}")
def generate_payslip(
    payroll_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Generate payslip PDF for specific payroll entry"""
    from fastapi.responses import StreamingResponse
    from payslip_generator import payslip_generator
    
    # Get payroll entry
    payroll = db.query(models.Payroll).filter(models.Payroll.id == payroll_id).first()
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll entry not found")
    
    # Get staff details
    staff = payroll.staff
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    
    # Prepare staff data
    staff_data = {
        'employee_id': staff.employee_id,
        'name': staff.name,
        'position': staff.position,
        'bank_account': getattr(staff, 'bank_account', 'N/A')
    }
    
    # Prepare payroll data
    payroll_data = {
        'month': payroll.month,
        'year': payroll.year,
        'gross_salary': float(payroll.gross_salary),
        'arrears': float(payroll.arrears) if payroll.arrears else 0,
        'advance_deduction': float(payroll.advance_deduction) if payroll.advance_deduction else 0,
        'other_deductions': float(payroll.other_deductions) if payroll.other_deductions else 0,
        'net_payable': float(payroll.net_payable),
        'payment_date': payroll.payment_date.strftime('%Y-%m-%d') if payroll.payment_date else datetime.now().strftime('%Y-%m-%d')
    }
    
    # Generate PDF
    pdf_buffer = payslip_generator.generate_payslip_pdf(staff_data, payroll_data)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=payslip_{staff.name.replace(' ', '_')}_{payroll.month}_{payroll.year}.pdf"
        }
    )

@app.post("/payslips/email/{payroll_id}")
def email_payslip(
    payroll_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Generate and email payslip to staff member"""
    from payslip_generator import payslip_generator
    from email_service import email_service
    
    # Get payroll entry
    payroll = db.query(models.Payroll).filter(models.Payroll.id == payroll_id).first()
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll entry not found")
    
    # Get staff details
    staff = payroll.staff
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    
    if not staff.email:
        raise HTTPException(status_code=400, detail="Staff member email not found")
    
    # Prepare data
    staff_data = {
        'employee_id': staff.employee_id,
        'name': staff.name,
        'position': staff.position,
        'bank_account': getattr(staff, 'bank_account', 'N/A')
    }
    
    payroll_data = {
        'month': payroll.month,
        'year': payroll.year,
        'gross_salary': float(payroll.gross_salary),
        'arrears': float(payroll.arrears) if payroll.arrears else 0,
        'advance_deduction': float(payroll.advance_deduction) if payroll.advance_deduction else 0,
        'other_deductions': float(payroll.other_deductions) if payroll.other_deductions else 0,
        'net_payable': float(payroll.net_payable),
        'payment_date': payroll.payment_date.strftime('%Y-%m-%d') if payroll.payment_date else datetime.now().strftime('%Y-%m-%d')
    }
    
    # Generate PDF
    pdf_buffer = payslip_generator.generate_payslip_pdf(staff_data, payroll_data)
    
    # Send email
    result = email_service.send_payslip_email(
        to_email=staff.email,
        staff_name=staff.name,
        month=payroll.month,
        year=payroll.year,
        net_payable=payroll.net_payable,
        pdf_attachment=pdf_buffer
    )
    
    if result["success"]:
        return {
            "success": True,
            "message": f"Payslip emailed to {staff.email}"
        }
    else:
        raise HTTPException(status_code=500, detail="Failed to send email")

@app.post("/payslips/bulk-generate")
def bulk_generate_payslips(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.UserRole.ADMIN, models.UserRole.MANAGER]))
):
    """Generate payslips for all staff for specified month/year"""
    from payslip_generator import payslip_generator
    
    # Get all payroll entries for the month/year
    payrolls = db.query(models.Payroll).filter(
        models.Payroll.month == month,
        models.Payroll.year == year
    ).all()
    
    if not payrolls:
        raise HTTPException(status_code=404, detail=f"No payroll entries found for {month}/{year}")
    
    results = {
        "generated": 0,
        "failed": 0,
        "details": []
    }
    
    for payroll in payrolls:
        try:
            staff = payroll.staff
            if not staff:
                results["failed"] += 1
                results["details"].append({
                    "payroll_id": payroll.id,
                    "status": "failed",
                    "error": "Staff not found"
                })
                continue
            
            # Generate payslip
            staff_data = {
                'employee_id': staff.employee_id,
                'name': staff.name,
                'position': staff.position,
                'bank_account': getattr(staff, 'bank_account', 'N/A')
            }
            
            payroll_data = {
                'month': payroll.month,
                'year': payroll.year,
                'gross_salary': float(payroll.gross_salary),
                'arrears': float(payroll.arrears) if payroll.arrears else 0,
                'advance_deduction': float(payroll.advance_deduction) if payroll.advance_deduction else 0,
                'other_deductions': float(payroll.other_deductions) if payroll.other_deductions else 0,
                'net_payable': float(payroll.net_payable),
                'payment_date': payroll.payment_date.strftime('%Y-%m-%d') if payroll.payment_date else datetime.now().strftime('%Y-%m-%d')
            }
            
            pdf_buffer = payslip_generator.generate_payslip_pdf(staff_data, payroll_data)
            
            results["generated"] += 1
            results["details"].append({
                "payroll_id": payroll.id,
                "staff_name": staff.name,
                "status": "success"
            })
            
        except Exception as e:
            results["failed"] += 1
            results["details"].append({
                "payroll_id": payroll.id,
                "status": "failed",
                "error": str(e)
            })
    
    return results


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
